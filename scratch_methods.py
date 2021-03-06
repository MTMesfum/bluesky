import pandas as pd, os, datetime, random, re, pickle
import numpy as np, timeit, shutil, warnings, math
import seaborn as sns; sns.set()
import matplotlib.pyplot as plt
from string import ascii_uppercase as alphabet
from bluesky.tools import aero
from bluesky.tools.geo import qdrdist as dist
from bluesky.tools.geo import latlondist as dist2
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from matplotlib.ticker import (MultipleLocator, FormatStrFormatter,
                               AutoMinorLocator, MaxNLocator)
import matplotlib.gridspec as gridspec
from matplotlib import rcParams

'''                 ######################################## 
                    ########################################
                    ###         Initial Settings!!!      ###
                    ########################################   
                    ########################################             
'''

settings_config = "settings.cfg"
scenario_manager = "scenario\Trajectories-batch3.scn"
scenario_manager2 = "scenario\Trajectories-batch4.scn"
exp = "scenario\Trajectories-batch2.scn"
save_ic = "scenario\\trajectories_saveic.scn"
dest_dir_input_logs = "output\\runs\\xlogs input"
dest_dir_output_logs = "output\\runs\\xlogs output"
dest_output = "output\\runs"
writer_file = 'output\WRITER Standard File.xlsx'
wind_ensemble = "Tigge_2014-09-08_00_243036.nc"
flight_date = "9,9,2014"
cruise_file = 'queries\AC Cruise Speed.xlsx'
cruise_file2 = 'queries\AC Cruise Speed 2.xlsx'
dt = 0.5
TW_place = True # position of the TW. True is in middle, False is on the bottom
TW_inf = 3600  # [s]
TW_min = 60  # [s]
set_of_delays = [0, 90, 720, 1050]  # [s]

pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)
warnings.filterwarnings("ignore")

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    UWARNING = '\033[4m' + '\033[93m'
    UBLUE = '\033[4m' + '\033[94m'


'''                 ########################################
                    ########################################
                    ###         Simulator controls!!!    ###
                    ########################################   
                    ########################################             
'''
def set_delays(input= [0, 90, 720, 1050]):
    global set_of_delays
    set_of_delays = input
    pass

def find_index(to_search, target):
    for number, i in enumerate(to_search):
        if target == i:
            return number
    return

# Replace the dt in the settings.cfg
def find_dt():

    with open(settings_config, 'r') as f:
        filedata = f.read()

    apple = filedata.find('simdt =')
    banana = filedata.find('# Snaplog dt')
    dt_settings = filedata[apple+8:banana-3]

    return dt_settings

# Switches the ensemble in the scenario manager and adapts the name using the ensemble # and global dt
def replace_ensemble(ensemble, file=None):
    if file == None:
        file = scenario_manager
    else:
        file = 'scenario\\' + file

    with open(file, 'r') as f:
        filedata = f.read()

    ensemble = str(ensemble)
    apple = filedata.find('LOAD_WIND')
    banana = filedata.find(', Tigge_')
    # citrus = filedata.find('WRITER')
    # date = filedata.find('_dt_')
    filedata = str("".join(filedata[0:apple+11] + str(ensemble).zfill(2) +
                           filedata[banana:]))
                           # filedata[banana:citrus+8] + str(ensemble) + filedata[date:]))

    with open(file, 'w') as f:
        f.write(filedata)

    # os.startfile(scenario_manager)
    pass

def replace_batch(scen_new):
    with open(scenario_manager, 'r') as f:
        filedata = f.read()

    apple = filedata.find('PCALL "')
    banana = filedata.find('.scn')
    filedata = str("".join(filedata[0:apple+7] + str(scen_new) +
                           filedata[banana+4:]))

    with open(scenario_manager, 'w') as f:
        f.write(filedata)

    # os.startfile(scenario_manager)

def replace_batch_set(scen_new, source, save_file):
    dir = os.getcwd()
    dest_dir = dir + "\\scenario\\" + save_file + '.scn'
    with open("scenario\\" + source + '.scn', 'r') as f:
        filedata = f.read()
    for i, j in enumerate(set_of_delays):
        apple = filedata.find('"Trajectories.scn"')
        banana = apple + len('"Trajectories.scn" ')
        filedata = str("".join(filedata[:apple+1] + str(scen_new)[:-4-len(str(set_of_delays[0]))]
                            + str(set_of_delays[i]) + '.scn" ' + str(i).zfill(2)
                               + filedata[banana+2:]))
        with open(dest_dir, 'w') as f:
            f.write(filedata)
        with open("scenario\\" + save_file + '.scn', 'r') as f:
            filedata = f.read()
    # os.startfile(dest_dir)

def replace_batch_set2(scen_new, source, save_file):
    dir = os.getcwd()
    folder = "\\scenario\\remon scen\\" + scen_new
    FileName = os.listdir(dir + folder)
    dest_dir = dir + "\\scenario\\" + save_file + '.scn'
    with open("scenario\\" + source + '.scn', 'r') as f:
        filedata = f.read()
    for i, j in enumerate(FileName):
        apple = filedata.find('"Trajectories.scn"')
        banana = apple + len('"Trajectories.scn" 00')
        filedata = str("".join(filedata[:apple+1] + "remon scen\\"
                               + str(scen_new) + "\\" + j + filedata[banana:]))
        with open(dest_dir, 'w') as f:
            f.write(filedata)
        with open("scenario\\" + save_file + '.scn', 'r') as f:
            filedata = f.read()

# Changes the timestep in the settings config of BlueSky using the provided timestep
# Keep in mind that the savefile doesn't change its name, unless the timestep is set into the global variable dt
def set_dt(timestep=dt):
    global dt
    dt = timestep

    with open(settings_config, 'r') as f:
        filedata = f.read()

    apple = filedata.find('simdt =')
    banana = filedata.find('# Performance')
    filedata = str("".join(filedata[0:apple+8] + str(timestep) + '\n \n' + filedata[banana:]))

    with open(settings_config, 'w') as f:
        f.write(filedata)

    print('\nThe timestep has been changed to {0} seconds.\n'.format(dt))
    # print('The dt is {0} seconds.\n'.format(dt))
    # os.startfile(settings_config)
    pass

# Switches the speed in run_experiment
def replace_speed(speed):
    with open(exp, 'r') as f:
        filedata = f.read()

    apple = filedata.find('run_experiment') #+16
    banana = filedata.find('00:00:00.00> FF')
    filedata = str("".join(filedata[0:apple+16] + str(speed) + filedata[banana-1:]))

    with open(exp, 'w') as f:
        f.write(filedata)

    # os.startfile(exp)
    pass

# Place the TW either in the middle or on the bottom
def set_TW_place(set):
    global TW_place
    TW_place = set
    pass

'''                 ######################################## 
                    ########################################
                    ###         Simulator Starters!!!    ###
                    ########################################   
                    ########################################             
'''
# Run a simulation of BlueSky using the desktop path
def bs_desktop():
    global drive_folder, drive_folder_inf, drive_folder_prob, drive_folder_det, \
        drive_folder_min, drive_folder_output, drive_folder_input, drive_folder_gen
    drive_folder = '/BlueSky Simulation/Run Desktop Home/'
    drive_folder_inf = '12LqgB2NtdHGU2EspFY5tgmO88fc53XEj'
    drive_folder_prob = '1N-V2AQzv2SnaUt_sGb6sxns3PZ3sEwmz'
    drive_folder_det = '1GzpfVUNmcCiM69mfvf0c-2eqCKIDqlYB'
    drive_folder_min = '1mE8bK9LmptxXHQbzKkSwSaNZsFw-HGGE'
    drive_folder_output = '1W_v0EuNL6WxpT18aHsf_PO6oHYXDDSNi'
    drive_folder_input = '1N7p25hQTLqmRUR_YAOPsdHvLHBeiRWF3'
    drive_folder_gen = '1p_OztE3UG3Oxed2UjEtmcV5FnMqwRPMf'
    os.system("call C:\Programs\Tools\Anaconda\Program\Scripts\\activate.bat && \
                    cd C:\Documents\Git 2 && conda activate py36 && python BlueSky.py")

# Run a simulation of BlueSky using the laptop path
def bs_laptop():
    global drive_folder, drive_folder_inf, drive_folder_prob, drive_folder_det, \
        drive_folder_min, drive_folder_output, drive_folder_input, drive_folder_gen
    drive_folder = '/BlueSky Simulation/Run Laptop/'
    drive_folder_inf = '1TKtpcnxWSrDTKG-7abW1joSuOCyXIaWi'
    drive_folder_prob = '1Cnum34BNGu3zsgoBMzugPPAhGUPV0R2i'
    drive_folder_det = '1G_v7JfZCyI3rs3X4bd2DH56bLK-QIIYM'
    drive_folder_min = '1FupY3LhocwprYZoz_bVs_rCy9xHZ4ljM'
    drive_folder_output = '1lN_lRg80jnKPzprTVE8BoKzSyeCIEaFk'
    drive_folder_input = '1pF06R0PXILh5DPQB0Rhd7sZ6CpS45-Ik'
    drive_folder_gen = '13PqsPnCCmMY_kruN95C38Fyn6uI2tCyY'
    os.system("call I:\Programs\Anaconda\Program\Scripts\\activate.bat && \
                    cd I:\Documents\Google Drive\Thesis 2018\BlueSky Git4 && python BlueSky.py")

# Run a simulation of BlueSky using the desktop TU 1 path
def bs_desktop_TU_1():
    global drive_folder, drive_folder_inf, drive_folder_prob, drive_folder_det, \
        drive_folder_min, drive_folder_output, drive_folder_input, drive_folder_gen
    drive_folder = '/BlueSky Simulation/Run Desktop TU 1/'
    drive_folder_inf = '1bqWP5kFObkz5_18e7Yu-qYdYWrGGRPn1'
    drive_folder_prob = '1LF4K_kaqsM77T5DRBLVWLgRiSR3LtR-W'
    drive_folder_det = '1d3hZ_IN_yxeeVGnfh3yrp75v1lJtWOXG'
    drive_folder_min = '1Tiho0ReAPGGMFwxYbEOhItLQSXfzGzHt'
    drive_folder_output = '1QFHeuLJ8APgsZru2PFpA6SXC2BShfZvl'
    drive_folder_input = '1b5fi3qI1fv0KuPJdyB3CFcKz_dQFHIYv'
    drive_folder_gen = '1f3WN6BuDU__JF_C6nQ2GE0C8ECxZyy2X'
    # use json1 + find path of anaconda!
    os.system("call D:\mmesfum\Programs\Anaconda\Scripts\\activate.bat && \
                    cd /d D:\mmesfum\Documents\BlueSky && conda activate py36 && python BlueSky.py")

# Run a simulation of BlueSky using the desktop TU 2 path
def bs_desktop_TU_2():
    global drive_folder, drive_folder_inf, drive_folder_prob, drive_folder_det, \
        drive_folder_min, drive_folder_output, drive_folder_input, drive_folder_gen
    drive_folder = '/BlueSky Simulation/Run Desktop TU 2/'
    drive_folder_inf = '1lqno4SXLh7o6BJ55xUYCyILiHViH9i5V'
    drive_folder_prob = '1QoSWj_d79El61sKn5lDNQ0MrsXMW4buP'
    drive_folder_det = '1kWcAOhbQpE1m-Ay-rdHnV2geIO6ynae2'
    drive_folder_min = '1rIpTDfhVIyKIuxpDLq4SMFXP3JvIaxNl'
    drive_folder_output = '1VWd6o5hTISw0ir5HKPLAXIZDu8vDzLO5'
    drive_folder_input = '1K_GjK_McE1KMg7YenogiPg4bAIOmKb-H'
    drive_folder_gen = '1RzqBSLCInHhzgkkkZ0bXOTrCjYih1rfe'
    # use json2 + find path of anaconda!
    os.system("call D:\mmesfum\Programs\Anaconda\Scripts\\activate.bat && \
                    cd /d D:\mmesfum\Documents\BlueSky && conda activate py36 && python BlueSky.py")


'''                 ######################################## 
                    ########################################
                    ###         Display Methods!!!       ###
                    ########################################   
                    ########################################             
'''

# Cut the number to 3 digits
def cut3(one):
    return round(one, 3)

# Cut the number to 7 digits
def cut7(one):
    return round(one, 7)

# Add a random delay to the time provided
# def addSecs(tm, secs, secs2):
def addSecs(time, set_of_delays2):
    if set_of_delays2 == []:
        set_of_delays2 = [0]
    list_of_times = [time + datetime.timedelta(seconds=x) for x in set_of_delays2]
    return list_of_times, set_of_delays2


'''                 ######################################## 
                    ########################################
                    ###         Scenario Creators!!!     ###
                    ########################################     
                    ########################################            
'''
# Create a scenario file from the provided trajectories and save it in the provided name
# The first entry in the method decides whether a random delay is added
def CreateSCN(alpha, save_file, folder = "queries\\"):
    FileName = os.listdir(folder)
    # FileName.remove('hide')
    traj = 0
    banana = list()
    scenario2 = pd.DataFrame()

    for k in FileName:
        _, ext = os.path.splitext(k)
        if ext != '.csv':
            continue
        if not scenario2.empty:
            scenario2 = scenario2.append(pd.read_csv(folder + "\\" + k))
        else:
            scenario2 = pd.read_csv(folder + "\\" + k)

    scenario2 = scenario2.sort_values(by=['time_over']).reset_index(drop=True)

    for l in set(scenario2['trajectory_id']):
        scenario    = scenario2[scenario2['trajectory_id'][:] == scenario2['trajectory_id'][0]].reset_index(drop=True)
        scenario2   = scenario2[scenario2['trajectory_id'][:] != scenario2['trajectory_id'][0]].reset_index(drop=True)

        traj += 1
        traj2 = str(traj)
        if traj2.count('') < 3:
            id = '0' + str(traj)
        else:
            id = str(traj)
        aircraftid = 'KLM' + id + '-%0'
        actype  =  ', A320, '
        [heading, distance] = dist(scenario['st_x(gpt.coords)'][0], scenario['st_y(gpt.coords)'][0],
                                    scenario['st_x(gpt.coords)'][1], scenario['st_y(gpt.coords)'][1])
        if heading < 0:
            heading += 360

        #speed = '250'
        apple = scenario.time_over[0]

        if alpha:
            apple = datetime.datetime(100, 1, 1, int(apple[-8:-6]), int(apple[-5:-3]), int(apple[-2:]))
            apple, delay = addSecs(apple, random.randrange(16), random.randrange(1))
        else:
            apple = apple[-8:]
        j = 0

        for i in range(scenario.shape[0]):
            if i != 0:
                j = -1

            # FlightLevel = str(scenario['fl'][i])
            # if FlightLevel.count('') < 3:
            #     FlightLevel = '0' + FlightLevel
            FlightLevel = 'FL' + str(scenario['fl'][i]).zfill(3)

            if i == 0:
                deltatime = datetime.datetime.strptime(scenario.time_over[i + 1], '%Y-%m-%d %H:%M:%S') - \
                            datetime.datetime.strptime(scenario.time_over[i ], '%Y-%m-%d %H:%M:%S')

                distance = abs(dist2(scenario['st_x(gpt.coords)'][i + 1], scenario['st_y(gpt.coords)'][i + 1],
                                     scenario['st_x(gpt.coords)'][i], scenario['st_y(gpt.coords)'][i]))

                height = abs((scenario['fl'][i + 1] - scenario['fl'][i]) * 100 * aero.ft)
                distance = (height ** 2 + distance ** 2) ** (1 / 2)

                speed = distance / deltatime.total_seconds()
                speed = aero.tas2cas(speed, int(FlightLevel[2:]) * 100 * aero.ft) * 3600 / aero.nm

                banana.append(apple + '.00> CRE ' + aircraftid + actype + str(cut7(scenario['st_x(gpt.coords)'][i]))
                              + ', ' + str(cut7(scenario['st_y(gpt.coords)'][i])) + ', '
                              + str(cut3(heading)) + ', 10, ' + str(cut3(speed))) #+ FlightLevel
                banana.append(apple + '.00> DEFWPT ' + aircraftid + '-ORIG,'
                              + str(cut7(scenario['st_x(gpt.coords)'][i])) + ', '
                              + str(cut7(scenario['st_y(gpt.coords)'][i])))
                banana.append(apple + '.00> ORIG ' + aircraftid + ', ' + aircraftid + '-ORIG')
                banana.append(apple + '.00> DEFWPT ' + aircraftid + '-DEST, '
                              + str(cut7(scenario['st_x(gpt.coords)'][scenario.shape[0]-1]))
                              + ', ' + str(cut7(scenario['st_y(gpt.coords)'][scenario.shape[0]-1])))
                banana.append(apple + '.00> DEST ' + aircraftid + ', ' + aircraftid + '-DEST' )
                banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid + '-DEST' + ' FL00/0.0')
                # banana.append(apple + '.00> DEST ' + aircraftid + ', '
                #                                + str(cut7(scenario['st_x(gpt.coords)'][scenario.shape[0]-1]))
                #                                + ' ' + str(cut7(scenario['st_y(gpt.coords)'][scenario.shape[0]-1])) )
            else:
                if i == scenario.shape[0]-1:
                    follow = aircraftid + '-' + str(i-1)
                    banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                                  + aircraftid + "-DEST" + ', ' + FlightLevel + ', ' + '0')
                    banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid + '-' + str(i-1) +
                                        ' ' + FL_OLD + '/0')
                else:
                    deltatime   = datetime.datetime.strptime(scenario.time_over[i], '%Y-%m-%d %H:%M:%S') -\
                                    datetime.datetime.strptime(scenario.time_over[i+j], '%Y-%m-%d %H:%M:%S')

                    distance    = abs(dist2(scenario['st_x(gpt.coords)'][i+j], scenario['st_y(gpt.coords)'][i+j],
                                     scenario['st_x(gpt.coords)'][i], scenario['st_y(gpt.coords)'][i]))

                    height      = abs((scenario['fl'][i] - scenario['fl'][i+j]) * 100 * aero.ft)
                    distance    = (height**2 + distance**2)**(1/2)

                    speed = distance / deltatime.total_seconds()
                    speed = aero.tas2cas(speed, int(FlightLevel[2:]) * 100 * aero.ft) * 3600 / aero.nm

                    if i == 1:
                        follow = aircraftid + '-ORIG'
                    else:
                        follow = aircraftid + '-' + str(i-1)

                    banana.append(apple + '.00> DEFWPT ' + aircraftid + '-' + str(i) + ', '
                                  + str(cut7(scenario['st_x(gpt.coords)'][i])) + ', '
                                  + str(cut7(scenario['st_y(gpt.coords)'][i])))
                    banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                                        + aircraftid + '-' + str(i) + ', ' + FlightLevel + ', ' + str(cut3(speed)))
                    if i >= 2:
                        banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid + '-' + str(i-1) +
                                        ' ' + FL_OLD + '/' + str(cut3(speed)))
            FL_OLD = FlightLevel

    #save = pd.DataFrame(banana)
    dir = os.getcwd()
    with open(dir + '/scenario/' + save_file + '.scn', "w") as fin:
        fin.write('\n'.join(banana))
    os.startfile(dir + '/scenario/' + save_file + '.scn')

    # os.startfile("F:\Documents\Python Scripts\ThesisScript")

# alpha = with delay
target_speed = 272.8 #A320
sub_FL = 0
def CreateSCN_Cruise(alpha, fl_ref, cap=999):
    if os.path.exists("scenario\\remon scen"):
        shutil.rmtree("scenario\\remon scen")
    folder = "scenario\\remon_raw_scenario\\"
    type_file = "AC Type 2.xlsx"
    FileName = os.listdir(folder)

    for index, k in enumerate(FileName):
        if cap == index:
            fl_ref = fl_ref - 20
        # print('FileName is: ', k)
        # print('FL_ref is: ', fl_ref)
        # print('cap is {} with index {}'.format(cap, index))
        if cap == index-1:
            return
        acid, ext = os.path.splitext(k)
        if ext != '.csv':
            continue
        scenario = pd.read_csv(folder + k)
        aircraftid = acid + '-%0'
        actype_file = pd.read_excel(folder + type_file)
        actype = actype_file[actype_file['id'].str.contains(k[0:7])]['AC type'].reset_index(drop=True)[0] + ', '
        TW_det = actype_file[actype_file['id'].str.contains(k[0:7])]['TW Det'].reset_index(drop=True)[0]
        TW_stoch = actype_file[actype_file['id'].str.contains(k[0:7])]['TW Stoch'].reset_index(drop=True)[0]

        scenario = scenario[scenario['fl']>fl_ref].reset_index(drop=True)
        apple = scenario.time_over[0]
        [heading, _] = dist(scenario['st_x(gpt.coords)'][0], scenario['st_y(gpt.coords)'][0],
                                    scenario['st_x(gpt.coords)'][1], scenario['st_y(gpt.coords)'][1])
        if heading < 0:
            heading += 360

        apple = datetime.datetime(100, 1, 1, int(apple[-8:-6]), int(apple[-5:-3]), int(apple[-2:]))
        basket_of_apples, set_of_delays2 = addSecs(apple, set_of_delays*alpha)

        replacement = zip(basket_of_apples, set_of_delays2)

        for apple, delay in replacement:
            apple = str(apple.time())
            m = 0
            for l in ['min', 'det', 'prob', 'inf']:
                banana = list()
                m += 1
                for i in range(scenario.shape[0]):
                    FlightLevel = 'FL' + str(scenario['fl'][i]-sub_FL).zfill(3)

                    if i == 0:
                        banana.append('00:00:00.00> FF')
                        banana.append(apple + '.00> CRE ' + aircraftid + ', ' + actype + str(cut7(scenario['st_x(gpt.coords)'][i]))
                                      + ', ' + str(cut7(scenario['st_y(gpt.coords)'][i])) + ', '
                                      + str(cut3(heading)) + ', ' + FlightLevel + ', ' + str(target_speed))
                        banana.append(apple + '.00> ASAS OFF')
                        # banana.append(apple + '.00> PRINTER A delay of {0} seconds has been added.'.format(str(delay)))
                        banana.append(apple + '.00> DEFWPT ' + aircraftid + '-ORIG, '
                                      + str(cut7(scenario['st_x(gpt.coords)'][i])) + ', '
                                      + str(cut7(scenario['st_y(gpt.coords)'][i])))
                        banana.append(apple + '.00> ORIG ' + aircraftid + ', ' + aircraftid + '-ORIG')
                        banana.append(apple + '.00> DEFWPT ' + aircraftid + '-DEST '
                                      + str(cut7(scenario['st_x(gpt.coords)'][scenario.shape[0]-1]))
                                      + ', ' + str(cut7(scenario['st_y(gpt.coords)'][scenario.shape[0]-1]))
                                      + ', FLYOVER')
                        banana.append(apple + '.00> DEST ' + aircraftid + ', ' + aircraftid + '-DEST')
                        banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid
                                                        + '-DEST ' + 'FL' + str(scenario['fl'][-1:].reset_index(drop=True)[0]))
                    else:
                        if i == scenario.shape[0]-1:
                            follow = aircraftid + '-' + str(i-1)
                            follow2 = aircraftid + '-' + str(i)
                            banana.append(apple + '.00> DEFWPT ' + follow2 + ', '
                                          + str(cut7(scenario['st_x(gpt.coords)'][scenario.shape[0] - 1]))
                                          + ', ' + str(cut7(scenario['st_y(gpt.coords)'][scenario.shape[0] - 1]))
                                          + ', FLYOVER')
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                                          + follow2 + ', ' + FlightLevel)
                            banana.append(apple + '.00> ' + aircraftid + ' AT ' + follow +
                                                ' ' + FL_OLD)
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow2 + ' ADDWPT '
                                          + aircraftid + "-DEST, " + FlightLevel)
                            banana.append(apple + '.00> ' + aircraftid + ' AT ' + follow2 +
                                                ' ' + FL_OLD)
                            o = list(range(0, scenario.shape[0], 5))
                            o.append(scenario.shape[0])
                            o.pop(0)
                            citrus1, citrus2, citrus3 = ([], [], [])
                            q = 1
                            time_to_add = 0
                            time = scenario.time_over[0]
                            time = datetime.datetime(1, 1, 1, int(time[-8:-6]), int(time[-5:-3]), int(time[-2:]))

                            for n in o:
                                if n == o[0]:
                                    wp_0 = aircraftid + '-ORIG ' #+ str(1) + ' '

                                wp_1 = aircraftid + '-' + str(n-1) + ' '

                                for p in range(q-1, n-1):
                                    [_, distance_local] = dist(scenario['st_x(gpt.coords)'][q-1],
                                                                scenario['st_y(gpt.coords)'][q-1],
                                                                scenario['st_x(gpt.coords)'][q],
                                                                scenario['st_y(gpt.coords)'][q])
                                    time_to_add += time_required2(distance_local, scenario['fl'][q]-sub_FL, target_speed)
                                    # time_to_add += time_required(distance_local, scenario['fl'][q]-sub_FL, target_speed)
                                    # print('Q is {} and time_to_add is {}'.format(q-1, time_to_add))
                                    q += 1

                                if l == 'min':      secs = TW_min/2
                                elif l == 'det':    secs = TW_det/2 * 60
                                elif l == 'prob':   secs = TW_stoch/2 * 60
                                elif l == 'inf':    secs = TW_inf/2

                                # now = dt.datetime.now()
                                delta = datetime.timedelta(seconds=(secs+round(time_to_add)))
                                # time_to_add = 0
                                t = time.time()
                                # print(t)
                                # 12:39:11.039864

                                # print(()
                                # print((dt.datetime.combine(dt.date(1, 1, 1), t) + delta).time())

                                # time = (datetime.datetime.combine(datetime.date(1, 1, 1), t) + delta)
                                time2 = (datetime.datetime.combine(datetime.date(1, 1, 1), t) + delta).time()
                                # print(time2)
                                banana.append(apple + '.00> ' + aircraftid + ' RTA_AT ' + wp_1 + str(time2))
                                citrus1.append(apple + '.00> ' + aircraftid + ' TW_SIZE_AT ' + wp_1 + str(int(secs*2)))
                                citrus2.append(apple + '.00> ' + aircraftid + ' OWN_SPD_FROM ' + wp_0 + ' ' + str(target_speed))
                                citrus3.append(apple + '.00> ' + aircraftid + ' AFMS_FROM ' + wp_0 + 'tw')
                                wp_0 = wp_1
                                del secs
                            banana = banana + citrus1 + citrus2 + citrus3
                            banana.append(apple + '.00> ' + aircraftid + ' AFMS_FROM ' + wp_1 + 'off')
                            banana.append(apple + '.00> VNAV ' + aircraftid + ' ON')
                            banana.append(apple + '.00> LNAV ' + aircraftid + ' ON')
                            # banana.append(apple + '.00> dt 0.5')

                        else:
                            if i == 1:
                                follow = aircraftid + '-ORIG'
                            else:
                                follow = aircraftid + '-' + str(i-1)

                            banana.append(apple + '.00> DEFWPT ' + aircraftid + '-' + str(i) + ', '
                                          + str(cut7(scenario['st_x(gpt.coords)'][i])) + ', '
                                          + str(cut7(scenario['st_y(gpt.coords)'][i])))
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                                                + aircraftid + '-' + str(i) + ', ' + FlightLevel)
                            if i >= 2:
                                banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid + '-' + str(i-1)
                                              + ' ' + FL_OLD)
                            [_, distance] = dist(scenario['st_x(gpt.coords)'][i], scenario['st_y(gpt.coords)'][i],
                                                    scenario['st_x(gpt.coords)'][i+1], scenario['st_y(gpt.coords)'][i+1])

                            # Time error calculation between database and bada
                            # time1 = scenario.time_over[i-1]
                            # time2 = scenario.time_over[i]
                            # time1 = datetime.datetime(100, 1, 1, int(time1[-8:-6]), int(time1[-5:-3]), int(time1[-2:]))
                            # time2 = datetime.datetime(100, 1, 1, int(time2[-8:-6]), int(time2[-5:-3]), int(time2[-2:]))
                            # time3 = time2 - time1
                            # print('\nWaypoint #', i)
                            # print('time required according to database: ', time3.total_seconds())
                            # print('time required according to bada: ',
                            #       time_required(distance, int(FlightLevel[2:]), 0.782))

                    FL_OLD = FlightLevel

                    dir = os.getcwd()
                    dest_dir = dir + '/scenario/remon scen/' + str(m) + ' ' + l + '/'

                    if not os.path.isdir(dest_dir):
                        os.makedirs(dest_dir)
                    with open(dest_dir + l + ' ' + acid + " D" + str(delay).strip() + '.scn', "w") as fin:
                        fin.write('\n'.join(banana))

            if not alpha:
                break


    # with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
    #     print(scenario)

    # os.startfile("F:\Documents\Python Scripts\ThesisScript")
    return

# This function is meant to create a set of scenarios which runs every trajectory
# of 1 ensemble of a case out of [min, det, stoch, inf]
def CreateSCN_Cruise2(alpha, selection=None, cap=999):
    if os.path.exists("scenario\\remon scen"):
        shutil.rmtree("scenario\\remon scen")
    folder = "scenario\\remon_raw_scenario\\"
    if not selection:
        selection = list()
        for i in os.listdir(folder):
            selection.append(os.path.splitext(i)[0])
    type_file = "AC Type 2.xlsx"
    FileName = os.listdir(folder)
    FileName2 = reversed(FileName)
    cruise_speed = pd.read_excel(cruise_file)
    sub_FL = 0

    # How to build new scenarios:
    # -> get flight
    # -> run through min / det / stoch / inf
    # 	-> get TW size for respective min etc
    # 		-> run through set of delays
    # 		-> add to file
    # 	-> save respective min etc
    number = 0
    for index, k in enumerate(FileName2):
        acid, ext = os.path.splitext(k)
        if cap == index-1:
            return
        elif acid not in selection:
            continue
        elif ext != '.csv' or acid in list(cruise_speed['skip']):
            FileName.remove(k)
            continue
        print('Creating scenario for Flight {}!'.format(k))
        number += 1
        scenario = pd.read_csv(folder + k)
        actype_file = pd.read_excel(folder + type_file)
        actype = actype_file[actype_file['id'].str.contains(k[0:-4])]['AC type'].reset_index(drop=True)[0] + ', '
        TW_det = actype_file[actype_file['id'].str.contains(k[0:-4])]['TW Det'].reset_index(drop=True)[0]
        TW_stoch = actype_file[actype_file['id'].str.contains(k[0:-4])]['TW Stoch'].reset_index(drop=True)[0]

        if not acid in cruise_speed[['switch_10', 'switch_20', 'switch_30', 'switch_60']]:
            fl_cor = 0
        elif acid in cruise_speed['switch_10']:     fl_cor = 10
        elif acid in cruise_speed['switch_20']:     fl_cor = 20
        elif acid in cruise_speed['switch_30']:     fl_cor = 30
        elif acid in cruise_speed['switch_60']:     fl_cor = 60

        fl_ref = max(scenario['fl']) - 2 - fl_cor
        scenario = scenario[scenario['fl'] > fl_ref].reset_index(drop=True)
        apple = scenario.time_over[0]
        [heading, _] = dist(scenario['st_x(gpt.coords)'][0], scenario['st_y(gpt.coords)'][0],
                                    scenario['st_x(gpt.coords)'][1], scenario['st_y(gpt.coords)'][1])
        if heading < 0:
            heading += 360

        apple = datetime.datetime(100, 1, 1, int(apple[-8:-6]), int(apple[-5:-3]), int(apple[-2:]))
        basket_of_apples, set_of_delays2 = addSecs(apple, set_of_delays*alpha)
        m = 0

        for l in ['min', 'det', 'prob', 'inf']:
            banana = list()
            banana.append('00:00:00.00> FF')
            m += 1
            target_speed = cruise_speed[cruise_speed['AC Type'] == actype[:-2]]['FL' + str(scenario['fl'][0])].item()
            for apple, delay in zip(basket_of_apples, set_of_delays2):
                apple = str(apple.time())
                aircraftid = acid + '-' + l + '-' + str(delay)
                for i in range(scenario.shape[0]):
                    FlightLevel = 'FL' + str(scenario['fl'][i]-sub_FL).zfill(3)

                    if i == 0:
                        banana.append(apple + '.00> CRE ' + aircraftid + ', ' + actype + str(cut7(scenario['st_x(gpt.coords)'][i]))
                                      + ', ' + str(cut7(scenario['st_y(gpt.coords)'][i])) + ', '
                                      + str(cut3(heading)) + ', ' + FlightLevel + ', ' + str(target_speed))
                        banana.append(apple + '.00> ASAS OFF')
                        # banana.append(apple + '.00> PRINTER A delay of {0} seconds has been added.'.format(str(delay)))
                        banana.append(apple + '.00> DEFWPT ' + aircraftid + '-ORIG, '
                                      + str(cut7(scenario['st_x(gpt.coords)'][i])) + ', '
                                      + str(cut7(scenario['st_y(gpt.coords)'][i])))
                        banana.append(apple + '.00> ORIG ' + aircraftid + ', ' + aircraftid + '-ORIG')
                        banana.append(apple + '.00> DEFWPT ' + aircraftid + '-DEST '
                                      + str(cut7(scenario['st_x(gpt.coords)'][scenario.shape[0]-1]))
                                      + ', ' + str(cut7(scenario['st_y(gpt.coords)'][scenario.shape[0]-1]))
                                      + ', FLYOVER')
                        banana.append(apple + '.00> DEST ' + aircraftid + ', ' + aircraftid + '-DEST')
                        banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid
                                                        + '-DEST ' + 'FL' + str(scenario['fl'][-1:].reset_index(drop=True)[0]))
                    else:
                        if i == scenario.shape[0]-1:
                            follow = aircraftid + '-' + str(i-1)
                            follow2 = aircraftid + '-' + str(i)
                            banana.append(apple + '.00> DEFWPT ' + follow2 + ', '
                                          + str(cut7(scenario['st_x(gpt.coords)'][scenario.shape[0] - 1]))
                                          + ', ' + str(cut7(scenario['st_y(gpt.coords)'][scenario.shape[0] - 1]))
                                          + ', FLYOVER')
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                                          + follow2 + ', ' + FlightLevel)
                            banana.append(apple + '.00> ' + aircraftid + ' AT ' + follow
                                          + ' ' + FL_OLD)
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow2 + ' ADDWPT '
                                          + aircraftid + "-DEST, " + FlightLevel)
                            banana.append(apple + '.00> ' + aircraftid + ' AT ' + follow2
                                          + ' ' + FL_OLD)
                            o = list(range(1, scenario.shape[0], 1))
                            o.append(scenario.shape[0])
                            o.pop(0)
                            citrus1, citrus2, citrus3 = ([], [], [])
                            q = 1
                            time_to_add = 0
                            time = scenario.time_over[0]
                            time = datetime.datetime(1, 1, 1, int(time[-8:-6]), int(time[-5:-3]), int(time[-2:]))

                            for n in o:
                                if n == o[0]:
                                    wp_0 = aircraftid + '-ORIG '

                                wp_1 = aircraftid + '-' + str(n-1) + ' '
                                citrus2.append(apple + '.00> ' + aircraftid + ' OWN_SPD_FROM '
                                               + wp_0 + ' ' + str(target_speed))

                                target_speed0 = target_speed
                                target_speed = cruise_speed[cruise_speed['AC Type'] == actype[:-2]][
                                    'FL' + str(scenario['fl'][q])].item()
                                if np.isnan(target_speed):
                                    print('{} has no speed at FL{}!!'.format(actype, str(scenario['fl'][q].item())))
                                    exit()
                                [_, distance_local] = dist(scenario['st_x(gpt.coords)'][q-1],
                                                            scenario['st_y(gpt.coords)'][q-1],
                                                            scenario['st_x(gpt.coords)'][q],
                                                            scenario['st_y(gpt.coords)'][q])

                                time_to_add += time_required2(distance_local, scenario['fl'][q]-sub_FL, target_speed)
                                if target_speed0 != target_speed:
                                    citrus2.append(apple + '.00> ' + aircraftid + ' OWN_SPD_FROM '
                                                   + wp_0 + ' ' + str(target_speed))
                                q += 1

                                if l == 'min':      secs = TW_min/2
                                elif l == 'det':    secs = TW_det/2 * 60
                                elif l == 'prob':   secs = TW_stoch/2 * 60
                                elif l == 'inf':    secs = TW_inf/2

                                if TW_place:
                                    delta = datetime.timedelta(seconds=(round(time_to_add)))
                                else:
                                    delta = datetime.timedelta(seconds=(secs+round(time_to_add)))

                                t = time.time()
                                time2 = (datetime.datetime.combine(datetime.date(1, 1, 1), t) + delta).time()
                                banana.append(apple + '.00> ' + aircraftid + ' RTA_AT ' + wp_1 + str(time2))
                                citrus1.append(apple + '.00> ' + aircraftid + ' TW_SIZE_AT ' + wp_1 + str(int(secs*2)))
                                citrus3.append(apple + '.00> ' + aircraftid + ' AFMS_FROM ' + wp_0 + 'tw')
                                wp_0 = wp_1
                                del secs
                            banana = banana + citrus1 + citrus2 + citrus3
                            banana.append(apple + '.00> ' + aircraftid + ' AFMS_FROM ' + wp_1 + 'off')
                            banana.append(apple + '.00> VNAV ' + aircraftid + ' ON')
                            banana.append(apple + '.00> LNAV ' + aircraftid + ' ON')

                        else:
                            if i == 1:
                                follow = aircraftid + '-ORIG'
                            else:
                                follow = aircraftid + '-' + str(i-1)

                            banana.append(apple + '.00> DEFWPT ' + aircraftid + '-' + str(i) + ', '
                                          + str(cut7(scenario['st_x(gpt.coords)'][i])) + ', '
                                          + str(cut7(scenario['st_y(gpt.coords)'][i])))
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                                                + aircraftid + '-' + str(i) + ', ' + FlightLevel)
                            if i >= 2:
                                banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid + '-' + str(i-1)
                                              + ' ' + FL_OLD)
                            [_, distance] = dist(scenario['st_x(gpt.coords)'][i], scenario['st_y(gpt.coords)'][i],
                                                    scenario['st_x(gpt.coords)'][i+1], scenario['st_y(gpt.coords)'][i+1])

                            # Time error calculation between database and bada
                            # time1 = scenario.time_over[i-1]
                            # time2 = scenario.time_over[i]
                            # time1 = datetime.datetime(100, 1, 1, int(time1[-8:-6]), int(time1[-5:-3]), int(time1[-2:]))
                            # time2 = datetime.datetime(100, 1, 1, int(time2[-8:-6]), int(time2[-5:-3]), int(time2[-2:]))
                            # time3 = time2 - time1
                            # print('\nWaypoint #', i)
                            # print('time required according to database: ', time3.total_seconds())
                            # print('time required according to bada: ',
                            #       time_required(distance, int(FlightLevel[2:]), 0.782))

                    FL_OLD = FlightLevel

            dest_dir = os.getcwd() + '/scenario/remon scen/' + str(m) + ' ' + l + '/'
            if not os.path.isdir(dest_dir):
                os.makedirs(dest_dir)
            with open(dest_dir + l + ' ' + acid + '.scn', "w") as fin:
                fin.write('\n'.join(banana))

    with open(os.getcwd() + '/scenario/number_of_ac.txt', "w+") as fin:
        number = number * len(set_of_delays)
        print('\nNumber of aircraft per simulation: ', number)
        fin.write(str(number))
    return

# This function is meant to create a set of scenarios which runs every trajectory
# of 1 ensemble of a case out of [min, det, stoch, inf]
def CreateSCN_Cruise3(alpha, selection=None, cap=999):
    if os.path.exists("scenario\\remon scen"):
        shutil.rmtree("scenario\\remon scen")
    folder = "scenario\\custom scen\\"
    if not selection:
        selection = list()
        for i in os.listdir(folder):
            selection.append(os.path.splitext(i)[0])
    FileName = os.listdir(folder)
    FileName2 = reversed(FileName)
    cruise_speed = pd.read_excel(cruise_file2)
    sub_FL = 0

    # How to build new scenarios:
    # -> get flight
    # -> run through min / det / stoch / inf
    # 	-> get TW size for respective min etc
    # 		-> run through set of delays
    # 		-> add to file
    # 	-> save respective min etc
    number = 0
    for index, k in enumerate(FileName2):
        acid, ext = os.path.splitext(k)
        if cap == index-1:
            return
        elif acid not in selection:
            print('Skipped scenario for Flight {}'.format(acid))
            continue
        elif ext != '.xlsx' or acid in list(cruise_speed['skip']):
            FileName.remove(k)
            continue
        print('Creating scenario for Flight {}'.format(acid))
        number += 1
        scenario = pd.read_excel(folder + k)
        actype = scenario['AAC-type'][0] + ','
        TW_det = (scenario['negative'][0] + scenario['positive'][0])

        if not acid in str(cruise_speed[['switch_10', 'switch_20', 'switch_30', 'switch_40', 'switch_50', 'switch_60']]):
            fl_cor = 0
        elif acid in str(cruise_speed['switch_10']):     fl_cor = 10
        elif acid in str(cruise_speed['switch_20']):     fl_cor = 20
        elif acid in str(cruise_speed['switch_30']):     fl_cor = 30
        elif acid in str(cruise_speed['switch_40']):     fl_cor = 40
        elif acid in str(cruise_speed['switch_50']):     fl_cor = 50
        elif acid in str(cruise_speed['switch_60']):     fl_cor = 60

        fl_ref = max(scenario['fl']) - 2 - fl_cor
        scenario = scenario[scenario['fl'] > fl_ref].reset_index(drop=True)
        apple = scenario.time_over[0]
        [heading, _] = dist(scenario['st_x(gpt.coords)'][0], scenario['st_y(gpt.coords)'][0],
                                    scenario['st_x(gpt.coords)'][1], scenario['st_y(gpt.coords)'][1])
        if heading < 0:
            heading += 360

        apple = datetime.datetime(100, 1, 1, int(apple[-5:-3]), int(apple[-2:]), 0)
        basket_of_apples, set_of_delays2 = addSecs(apple, set_of_delays*alpha)
        m = 0

        for l in ['min', 'det', 'inf']:
            banana = list()
            banana.append('00:00:00.00> FF')
            m += 1

            for apple, delay in zip(basket_of_apples, set_of_delays2):
                apple = str(apple.time())
                aircraftid = acid + '-' + l + '-' + str(delay)
                for i in range(scenario.shape[0]):
                    FlightLevel = 'FL' + str(scenario['fl'][i]-sub_FL).zfill(3)
                    target_speed = cruise_speed[cruise_speed['AC Type'] == actype[:-1]][
                        'FL' + str(round(scenario['fl'][0], -1))].item()
                    if i == 0:
                        banana.append(apple + '.00> CRE ' + aircraftid + ', ' + actype + str(cut7(scenario['st_x(gpt.coords)'][i]))
                                      + ', ' + str(cut7(scenario['st_y(gpt.coords)'][i])) + ', '
                                      + str(cut3(heading)) + ', ' + FlightLevel + ', ' + str(target_speed))
                        banana.append(apple + '.00> ASAS OFF')
                        # banana.append(apple + '.00> PRINTER A delay of {0} seconds has been added.'.format(str(delay)))
                        banana.append(apple + '.00> DEFWPT ' + aircraftid + '-ORIG, '
                                      + str(cut7(scenario['st_x(gpt.coords)'][i])) + ', '
                                      + str(cut7(scenario['st_y(gpt.coords)'][i])))
                        banana.append(apple + '.00> ORIG ' + aircraftid + ', ' + aircraftid + '-ORIG')
                        banana.append(apple + '.00> DEFWPT ' + aircraftid + '-DEST '
                                      + str(cut7(scenario['st_x(gpt.coords)'][scenario.shape[0]-1]))
                                      + ', ' + str(cut7(scenario['st_y(gpt.coords)'][scenario.shape[0]-1]))
                                      + ', FLYOVER')
                        banana.append(apple + '.00> DEST ' + aircraftid + ', ' + aircraftid + '-DEST')
                        banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid
                                                        + '-DEST ' + 'FL' + str(scenario['fl'][-1:].reset_index(drop=True)[0]))
                    else:
                        if i == scenario.shape[0]-1:
                            follow = aircraftid + '-' + str(i-1)
                            follow2 = aircraftid + '-' + str(i)
                            banana.append(apple + '.00> DEFWPT ' + follow2 + ', '
                                          + str(cut7(scenario['st_x(gpt.coords)'][scenario.shape[0] - 1]))
                                          + ', ' + str(cut7(scenario['st_y(gpt.coords)'][scenario.shape[0] - 1]))
                                          + ', FLYOVER')
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                                          + follow2 + ', ' + FlightLevel)
                            banana.append(apple + '.00> ' + aircraftid + ' AT ' + follow
                                          + ' ' + FL_OLD)
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow2 + ' ADDWPT '
                                          + aircraftid + "-DEST, " + FlightLevel)
                            banana.append(apple + '.00> ' + aircraftid + ' AT ' + follow2
                                          + ' ' + FL_OLD)
                            o = list(range(1, scenario.shape[0], 1))
                            o.append(scenario.shape[0])
                            o.pop(0)
                            citrus1, citrus2, citrus3 = ([], [], [])
                            q = 1
                            time_to_add = 0
                            time = scenario.time_over[0]
                            time = datetime.datetime(1, 1, 1, int(time[-8:-6]), int(time[-5:-3]), int(time[-2:]))

                            for n in o:
                                if n == o[0]:
                                    wp_0 = aircraftid + '-ORIG '
                                elif n != o[-1]:
                                    citrus2.append(apple + '.00> ' + aircraftid + ' OWN_SPD_FROM '
                                                   + wp_0 + ' ' + str(target_speed))

                                wp_1 = aircraftid + '-' + str(n-1) + ' '

                                target_speed0 = target_speed
                                target_speed = cruise_speed[cruise_speed['AC Type'] == actype[:-1]][
                                    'FL' + str(round(scenario['fl'][q], -1))].item()
                                if np.isnan(target_speed):
                                    print('{} has no speed at FL{}!!'.format(actype, str(scenario['fl'][q].item())))
                                    exit()

                                time2 = scenario.time_over[q][-6:]
                                if target_speed0 != target_speed:
                                    citrus2.append(apple + '.00> ' + aircraftid + ' OWN_SPD_FROM '
                                                   + wp_0 + ' ' + str(target_speed))
                                q += 1

                                if l == 'min':      secs = TW_min/2
                                elif l == 'det':    secs = TW_det/2 * 60
                                elif l == 'inf':    secs = TW_inf/2

                                # Currently, this 'createscn' can not place the TW at the bottom, so middle only!
                                if TW_place:
                                    # time2 = ...
                                    delta = datetime.timedelta(seconds=(round(time_to_add)))
                                else:
                                    # time2 = ...
                                    delta = datetime.timedelta(seconds=(secs+round(time_to_add)))

                                # t = time.time()
                                # time2 = (datetime.datetime.combine(datetime.date(1, 1, 1), t) + delta).time()

                                banana.append(apple + '.00> ' + aircraftid + ' RTA_AT ' + wp_1 + str(time2) + ':00')
                                citrus1.append(apple + '.00> ' + aircraftid + ' TW_SIZE_AT ' + wp_1 + str(int(secs*2)))
                                citrus3.append(apple + '.00> ' + aircraftid + ' AFMS_FROM ' + wp_0 + 'tw')
                                wp_0 = wp_1
                                del secs
                            banana = banana + citrus1 + citrus2 + citrus3
                            banana.append(apple + '.00> ' + aircraftid + ' AFMS_FROM ' + wp_1 + 'off')
                            banana.append(apple + '.00> VNAV ' + aircraftid + ' ON')
                            banana.append(apple + '.00> LNAV ' + aircraftid + ' ON')

                        else:
                            if i == 1:
                                follow = aircraftid + '-ORIG'
                            else:
                                follow = aircraftid + '-' + str(i-1)

                            banana.append(apple + '.00> DEFWPT ' + aircraftid + '-' + str(i) + ', '
                                          + str(cut7(scenario['st_x(gpt.coords)'][i])) + ', '
                                          + str(cut7(scenario['st_y(gpt.coords)'][i])))
                            banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                                                + aircraftid + '-' + str(i) + ', ' + FlightLevel)
                            if i >= 2:
                                banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid + '-' + str(i-1)
                                              + ' ' + FL_OLD)
                            [_, distance] = dist(scenario['st_x(gpt.coords)'][i], scenario['st_y(gpt.coords)'][i],
                                                    scenario['st_x(gpt.coords)'][i+1], scenario['st_y(gpt.coords)'][i+1])

                            # Time error calculation between database and bada
                            # time1 = scenario.time_over[i-1]
                            # time2 = scenario.time_over[i]
                            # time1 = datetime.datetime(100, 1, 1, int(time1[-8:-6]), int(time1[-5:-3]), int(time1[-2:]))
                            # time2 = datetime.datetime(100, 1, 1, int(time2[-8:-6]), int(time2[-5:-3]), int(time2[-2:]))
                            # time3 = time2 - time1
                            # print('\nWaypoint #', i)
                            # print('time required according to database: ', time3.total_seconds())
                            # print('time required according to bada: ',
                            #       time_required(distance, int(FlightLevel[2:]), 0.782))

                    FL_OLD = FlightLevel

            dest_dir = 'scenario\\remon scen\\' + str(m) + ' ' + l + '\\'
            if not os.path.isdir(dest_dir):
                os.makedirs(dest_dir)
            with open(dest_dir + l + ' ' + acid + '.scn', "w") as fin:
                fin.write('\n'.join(banana))

    with open('scenario\\number_of_ac.txt', "w+") as fin:
        number = number * len(set_of_delays)
        print('\nNumber of aircraft per simulation: ', number)
        fin.write(str(number))
    return

def CreateSCN_FE(actype, FlightLevel, v_bada, delta, steps=0.001):
    # Setup for the SCN Manager and scenarios
    to_save = actype + ' ' + FlightLevel + '.scn'
    dir = 'data/performance/BS/aircraft'
    to_run = 'Flight Test/' + to_save
    actype_doc = actype + '.xml'
    path = os.path.join(dir, actype_doc)
    number = 0
    # start 50 3.5, [1] 51 3.5, [2] 52 3.5, [3] 53 3.5, [4] 54 3.5, [5] 55 3.5
    lat_start = 50
    lon_start = 3.5
    heading_start = 33
    # 51.75 2.17 35
    # lat_start = 51.75
    # lon_start = 2.17
    # heading_start = 19
    # WPT1: 54 3.5
    # Dest: 55 3.6

    # Create a SCN Manager
    files = os.listdir(dir)
    if actype_doc not in files:
        print('AC Type not found!')
        # os.startfile('C:\Documents\Git 2\\' + dir)
        # exit()
    # ac_doc = ElementTree.parse(path)
    # v_bada = float(ac_doc.find('speeds/cr_MA').text)

    gamma = list()
    gamma.append('# This is the scenario manager to find the speed corresponding to the lowest fuel consumption')
    gamma.append('00:00:00.00> SWRAD VOR')
    gamma.append('00:00:00.00> SAVEIC trajectories_saveic')
    gamma.append('00:00:00.00> CRELOG MYLOG 0.1 "MYLOG"')
    gamma.append('00:00:00.00> MYLOG ADD traf.id, traf.ax, traf.gs, pilot.tas, traf.lat, '
                 'traf.lon, traf.perf.fuelflow')
    gamma.append('00:00:00.00> MYLOG ON')
    gamma.append('00:00:00.00> FF')

    for i, j in enumerate(np.arange(v_bada - delta, v_bada + delta, steps)):
        number += 1
        gamma.append('')
        gamma.append('# Load trajectories for a run')
        gamma.append('00:00:00.00> SCEN Test_' + str(i).zfill(2))
        j2 = str("%.4f" % j)[2:]
        gamma.append('00:00:00.00> PCALL "{}" {}, {}'.format(to_run, str(i).zfill(2), str(j2)))
                     # + "%.4f" % j)
        gamma.append('23:59:00.00> HOLD')

    # Create set of scenarios
    aircraftid = actype + '-%0'
    banana = list()
    apple = '00:00:00'
    banana.append(apple + '.00> CRE ' + aircraftid + ', ' + actype + ', ' + str(lat_start) + ', '
                  + str(lon_start) + ', ' + str(heading_start) + ', ' + FlightLevel + ', 0.%1')
    # banana.append(apple + '.00> SPD ' + aircraftid + ', 0.%1')
    banana.append(apple + '.00> ASAS OFF')
    banana.append(apple + '.00> DEFWPT ' + aircraftid + '-ORIG '
                  + str(lat_start) + ', ' + str(lon_start) + ', FLYOVER')
    banana.append(apple + '.00> ORIG ' + aircraftid + ', ' + aircraftid + '-ORIG')
    banana.append(apple + '.00> DEFWPT ' + aircraftid + '-DEST '
                  + str(lat_start+5) + ', ' + str(lon_start+5) + ', FLYOVER')
    banana.append(apple + '.00> DEST ' + aircraftid + ', ' + aircraftid + '-DEST')
    banana.append(apple + '.00> ' + aircraftid + ' AT ' + aircraftid + '-DEST ' + FlightLevel)
    follow = aircraftid + '-ORIG'

    for k in range(1, 5):
        follow2 = aircraftid + '-' + str(k)
        banana.append(apple + '.00> DEFWPT ' + follow2 + ', ' + str(lat_start+k)
                      + ', ' + str(lon_start+k) + ', FLYOVER')
        banana.append(apple + '.00> ' + aircraftid + ' after ' + follow + ' ADDWPT '
                      + follow2 + ', ' + FlightLevel)
        follow = follow2
    banana.append('00:00:15.00> FF')
    banana.append('23:00:00.00> EXIT')

    # Save the scenarios
    dest_dir = os.getcwd() + '\scenario\Flight Test\\'
    if not os.path.isdir(dest_dir):
        os.makedirs(dest_dir)
    with open(dest_dir + to_save, "w") as fin:
        fin.write('\n'.join(banana))

    # Save the SCNM after adding the distances between the waypoints
    for k in range(1, 4):
        _, distance = dist(lat_start+k, lon_start+k, lat_start+k+1, lon_start+k+1)
        gamma.append('# Distance between Waypoint {} and {} is {}.'.format(k-1, k, round(distance, 2)))

    with open(dest_dir + 'SCNM_FE.scn', "w") as fin:
        print('SCNM_FE has been changed to type {} and {}!'.format(actype, FlightLevel))
        fin.write('\n'.join(gamma))

    with open(os.getcwd() + '/scenario/number_of_ac.txt', "w+") as fin:
        fin.write(str(number))
    pass

# Create a scenario manager to run a file alpha times, load wind ensemble beta and save the file in save_file
def CreateSCNM(alpha, beta, save_file):
    gamma = list()
    gamma.append('# Load wind data')
    gamma.append('00:00:00.00> LOAD_WIND ' + str(beta).zfill(2) + ', Tigge_01_09_2017.nc')
    gamma.append('00:00:00.00> DATE 1,9,2017')
    gamma.append('00:00:00.00> FF')

    for i in range(1, alpha):
        gamma.append('')
        gamma.append('# Load trajectories for a run')
        gamma.append('00:00:00.00> SWRAD VOR')
        gamma.append('00:00:00.00> SCEN Test_' + str(i).zfill(2))
        gamma.append('00:00:00.00> PCALL "Trajectories.scn" ' + str(i).zfill(2))
        gamma.append('00:00:00.00> FF')
        gamma.append('23:59:00.00> HOLD')

    gamma.append('')
    gamma.append('# Load trajectories for a run')
    gamma.append('00:00:00.00> SCEN Test_' + str(alpha).zfill(2))
    gamma.append('00:00:00.00> PCALL "Trajectories.scn" ' + str(alpha).zfill(2))
    gamma.append('00:00:00.00> SWRAD VOR')
    gamma.append('00:00:00.00> FF')
    gamma.append('23:58:59.00> WRITER W' + str(beta).zfill(2) + '_dt_' + str(dt))
    gamma.append('23:58:59.00> EXIT')

    dir = os.getcwd()
    with open(dir + '/scenario/' + save_file + '.scn', "w") as fin:
        fin.write('\n'.join(gamma))
    os.startfile(dir + '/scenario/' + save_file + '.scn')

# Create a different scenario manager to run a file [alpha times], load wind ensemble beta and save the file in save_file
def CreateSCNM2(save_file):
    alpha = len(set_of_delays)
    beta = 1
    gamma = list()
    gamma.append('# Load wind data')
    gamma.append('00:00:00.00> SAVEIC trajectories_saveic')
    gamma.append('00:00:00.00> CRELOG MYLOG 0.1 "MYLOG"')
    # gamma.append('00:00:00.00> MYLOG ADD traf.id, traf.lat, traf.lon, traf.perf.mass,')
    gamma.append('00:00:00.00> MYLOG ADD traf.id, traf.ax, traf.gs, pilot.tas, traf.lat, traf.lon, traf.perf.fuelflow')
    gamma.append('00:00:00.00> MYLOG ON')
    gamma.append('00:00:00.00> ASAS OFF')
    gamma.append('00:00:00.00> DATE ' + flight_date)
    gamma.append('00:00:00.00> LOAD_WIND2 ' + str(beta).zfill(2) +', ' + wind_ensemble)

    for i in range(1, alpha):
        gamma.append('')
        gamma.append('# Load trajectories for a run')
        gamma.append('00:00:00.00> SCEN Test_' + str(i).zfill(2))
        gamma.append('00:00:00.00> PCALL "Trajectories.scn" ' + str(i).zfill(2))
        gamma.append('00:00:00.00> FF')
        gamma.append('23:59:00.00> HOLD')

    gamma.append('')
    gamma.append('# Load trajectories for a run')
    gamma.append('00:00:00.00> SWRAD VOR')
    gamma.append('00:00:00.00> SCEN Test_' + str(alpha).zfill(2))
    gamma.append('00:00:00.00> PCALL "Trajectories.scn" ' + str(alpha).zfill(2))
    gamma.append('00:00:00.00> FF')
    # gamma.append('23:58:59.00> WRITER W' + str(beta).zfill(2) +'_dt_' +str(dt))
    gamma.append('23:58:59.00> EXIT')

    with open('scenario\\' + save_file + '.scn', "w") as fin:
        fin.write('\n'.join(gamma))
    # os.startfile('scenario\\' + save_file + '.scn')

# Create a different scenario manager to run a every traj of one ensemble
def CreateSCNM3(save_file, input="scenario\\remon scen\\1 min"):
    alpha = len(os.listdir(os.path.join(os.getcwd(), input)))
    beta = 1
    gamma = list()
    gamma.append('# Load wind data')
    gamma.append('00:00:00.00> SAVEIC trajectories_saveic')
    gamma.append('00:00:00.00> CRELOG MYLOG 10 "MYLOG"')
    # gamma.append('00:00:00.00> MYLOG ADD traf.id, traf.lat, traf.lon, traf.perf.mass,')
    gamma.append('00:00:00.00> MYLOG ADD traf.id, traf.ax, traf.gs, pilot.tas, traf.lat, traf.lon, traf.perf.fuelflow')
    gamma.append('00:00:00.00> MYLOG ON')
    gamma.append('00:00:00.00> SWRAD VOR')
    gamma.append('00:00:00.00> ASAS OFF')
    gamma.append('00:00:00.00> DATE ' + flight_date)
    gamma.append('00:00:00.00> LOAD_WIND2 ' + str(beta).zfill(2) +', ' + wind_ensemble)

    for i in range(1, alpha+1):
        gamma.append('')
        gamma.append('# Load trajectories for a run')
        gamma.append('00:00:00.00> SCEN Test_' + str(i).zfill(2))
        gamma.append('00:00:00.00> PCALL "Trajectories.scn" ' + str(i).zfill(2))
        gamma.append('00:00:00.00> FF')
        gamma.append('23:59:00.00> HOLD')

    # gamma.pop(-1)
    # gamma.append('23:58:59.00> WRITER W' + str(beta).zfill(2) +'_dt_' +str(dt))
    # gamma.append('23:58:59.00> EXIT')

    with open('scenario\\' + save_file + '.scn', "w") as fin:
        fin.write('\n'.join(gamma))
    # os.startfile('scenario\\' + save_file + '.scn')


'''                 ######################################## 
                    ########################################
                    ###   Simulation Save Controls!!!    ###
                    ########################################   
                    ########################################             
'''
# Clean up,  and open the output file
def writerfix(traj, dir, counter):
    dest_dir = 'output\\runs\{0}'.format(dir)
    df = pd.read_excel('output\WRITER Standard File.xlsx')
    # df = pd.read_csv('output\\runs\WRITER {0}.csv'.format(traj))
    # print(df)
    # df = df.drop('Unnamed: 0', axis=1)
    df.reset_index(inplace=True)
    df.index = df.index + 1
    df = df.drop('index', axis=1)
    name, _ = os.path.splitext(traj)

    dest_dir_input_logs2 = dest_dir_input_logs + "\\" + dir[0:7] + "\\" + traj[4:10]
    with open(dest_dir_input_logs2 + "\\" + dir[0:7] + " " + traj[4:10] +
              " IE01 D{}.scn".format(str(set_of_delays[0])), 'r') as f:
        filedata = f.read()
    apple = filedata.find('TW_SIZE_AT ')
    banana = filedata[apple+len('TW_SIZE_AT '):].find('TW_SIZE_AT ') + apple
    _, _, TW_size, *_ = filedata[apple:banana].split()

    if not os.path.isdir(dest_dir):
        os.mkdir(dest_dir)
    df.to_excel(dest_dir + '\{0} TW={1}.xlsx'.format(name, TW_size))
    os.remove('output\WRITER Standard File.xlsx')
    print(bcolors.UBLUE     + '\n\nSaved'   +
          bcolors.FAIL      + ' [{1}] {0} '.format(traj, counter)   +
          bcolors.UBLUE     + 'in'      +
          bcolors.FAIL      + ' {0}'.format(
                              dest_dir + '\{0}.xlsx'.format(name)) +
          bcolors.ENDC)
    # os.startfile('output\\runs\WRITER {0}.xlsx'.format(traj))

# Clean up,  and open the output file
def writerfix2(dir, counter, upload):
    dest_dir = 'output\\runs\{0}'.format(dir)
    with open(scenario_manager2, 'r') as f:
        filedata = f.read()
    apple = filedata.find('LOAD_WIND2')
    banana = filedata.find(', ' + wind_ensemble)
    # [apple+len('TW_SIZE_AT '):].find('TW_SIZE_AT ') + apple
    # _, _, TW_size, *_ = filedata[apple:banana].split()
    ensemble = str(filedata[apple+11:banana]).zfill(2)

    # df.to_excel(dest_dir + '\{0} TW={1}.xlsx'.format(name, TW_size))
    new_name0 = '{} Ensemble={}.xlsx'.format(dir, ensemble)
    new_name = dest_dir + '\\{} Ensemble={}.xlsx'.format(dir, ensemble)
    # print('new name is: ', new_name)
    if not os.path.isdir(dest_dir):
        os.mkdir(dest_dir)
    os.rename(writer_file, new_name)
    if upload:
        drive_folder2 = drive_folder + dir + '/'
        upload_file(new_name, new_name0, drive_folder2)

    print(bcolors.UBLUE     + '\n\nSaved'   +
          bcolors.FAIL      + ' [{1}] {0} '.format(dir, counter)   +
          bcolors.UBLUE     + 'in'      +
          bcolors.FAIL      + ' {0}'.format(new_name) +
          bcolors.ENDC)
    # os.startfile('output\\runs\WRITER {0}.xlsx'.format(traj))

# 'movelog' has been replaced by movelog2 and "PRINTER" lines have been disabled
def movelog(ensemble, traj, dir):
    #Move the input log file into the input log folder
    dest_dir_input_logs2 = dest_dir_input_logs + "\\" + dir[0:7] + "\\" + traj[4:10]
    if not os.path.isdir(dest_dir_input_logs2):
        os.makedirs(dest_dir_input_logs2)

    with open(save_ic, 'r') as f:
        filedata = f.read()

    apple = filedata.find('PRINTER A delay of') #+16
    banana = filedata.find('seconds has been added')
    delay = filedata[apple+len('PRINTER A delay of '):banana]

    new_name = "\\" + dir[0:7] + " " + traj[4:10] + \
               " IE" + str(ensemble).zfill(2) + " D" + str(delay).strip() + ".scn"
    os.rename(save_ic, dest_dir_input_logs2 + new_name)

    #Move the output log file into the output log folder if the file exists
    dest_dir_output_logs2 = dest_dir_output_logs + "\\" + dir[0:7] + "\\" + traj[4:10]
    if not os.path.isdir(dest_dir_output_logs2):
        os.makedirs(dest_dir_output_logs2)
    files = os.listdir('output\\')
    files = [files[files.index(i)] for i in files if 'MYLOG' in i]
    if files == []:
        pass
    else:
        files = str(files[-1])
        new_name = "\\" + dir[0:7] + " " + traj[4:-4] + \
                   " OE" + str(ensemble).zfill(2) + " D" + str(delay).strip() + ".log"
        os.rename("output\\" + files, dest_dir_output_logs2 + new_name)
        # Analyse the data and open the file
        # compare_ff(dest_dir_output_logs2 + new_name)
    pass

def movelog2(ensemble_int, dir, upload):
    #Move the input log file into the input log folder
    dest_dir_input_logs2 = dest_dir_input_logs + "\\" + dir[0:7]
    if not os.path.isdir(dest_dir_input_logs2):
        os.makedirs(dest_dir_input_logs2)

    with open(save_ic, 'r') as f:
        filedata = f.read()

    apple = filedata.find('LOAD_WIND2') #+16
    banana = filedata.find(', ' + wind_ensemble)
    ensemble = str(filedata[apple+11:banana]).zfill(2)

    if ensemble != str(ensemble_int).zfill(2):
        print("The ensembles don't match!")
        print("Ensemble should be {} but {} has been read in!".format(ensemble_int, ensemble))

    new_name = '{} IEnsemble={}.scn'.format(dir, ensemble)
    new_name2 = dest_dir_input_logs2 + "\\" + new_name
    os.rename(save_ic, new_name2)
    if upload:
        drive_folder2 = drive_folder + dest_dir_input_logs[-11:] + '/'
        upload_file(new_name2, new_name, drive_folder2)

    #Move the output log file into the output log folder if the file exists
    dest_dir_output_logs2 = dest_dir_output_logs + "\\" + dir[0:7]
    if not os.path.isdir(dest_dir_output_logs2):
        os.makedirs(dest_dir_output_logs2)
    files = os.listdir('output\\')
    files = [files[files.index(i)] for i in files if 'MYLOG' in i]
    if files == []:
        pass
    else:
        new_name = '{} OEnsemble={}.log'.format(dir, ensemble)
        new_name2 = dest_dir_output_logs2 + '\\' + new_name
        os.rename("output\\" + str(files[-1]), new_name2)
        if upload:
            drive_folder2 = drive_folder + dest_dir_output_logs[-12:] + '/'
            upload_file(new_name2, new_name, drive_folder2)
        # Analyse the data and open the file
        # compare_ff(dest_dir_output_logs2 + new_name)
    pass

def clear_mylog():
    files = os.listdir('output\\')
    files = [files[files.index(i)] for i in files if 'MYLOG' in i]
    for i in files:
        os.remove('output\\' + i)
        print('The file {} has been deleted!'.format(i))

'''                 ######################################## 
                    ########################################
                    ###          Timer Controls!!!       ###
                    ########################################   
                    ########################################             
'''
def talk_time(runs):
    print(bcolors.UWARNING + "\nExecuting simulation {1} and {0} seconds have passed!".format(
        int(timeit.default_timer()), runs) + bcolors.ENDC)
    if runs > 1:
        print(bcolors.UWARNING +
              "This results in {0} seconds per run!".format(int(timeit.default_timer() / (runs - 1)))
              + bcolors.ENDC)

def talk_traj(scen_next, traj_counter):
    print(bcolors.UWARNING + '\nReplaced Trajectory to' +
          bcolors.FAIL + ' [{1}] {0}'.format(scen_next, traj_counter + 1) +
          bcolors.ENDC)

def talk_traj2(scen_next, traj_counter):
    print(bcolors.UWARNING + '\nReplaced Trajectories to' +
          bcolors.FAIL + ' [{1}] {0}'.format(scen_next[:-7], traj_counter + 1) +
          bcolors.ENDC)

def talk_run(ensemble, sgl_traj, traj_counter, dir):
    print(bcolors.UWARNING + '\nRunning Trajectory' +
          bcolors.FAIL + ' [{1}] {2}\{0} '.format(sgl_traj, traj_counter + 1, dir) +
          bcolors.UWARNING + 'with Ensemble' +
          bcolors.FAIL + ' [{0}]\n'.format(str(ensemble).zfill(2)) +
          bcolors.ENDC)

def talk_run2(ensemble, sgl_traj, traj_counter, dir):
    print(bcolors.UWARNING + '\nRunning Trajectory' +
          bcolors.FAIL + ' [{1}] {2}\{0} '.format(sgl_traj[:-7], traj_counter + 1, dir) +
          bcolors.UWARNING + 'with Ensemble' +
          bcolors.FAIL + ' [{0}]\n'.format(str(ensemble).zfill(2)) +
          bcolors.ENDC)

def talk_run3(ensemble, dir, runs):
    print(bcolors.UWARNING + '\nRunning Trajectory directory' +
          bcolors.FAIL + ' {} '.format(dir, runs) +
          bcolors.UWARNING + 'with Run / Ensemble:' +
          bcolors.FAIL + ' [{} / {}]\n'.format(runs, str(ensemble).zfill(2)) +
          bcolors.ENDC)

def time_required(distance, FL, speed):
    # distance in nm and speed in mach
    tas = aero.vcas2tas(speed * aero.nm / 3600, FL * 100 * aero.ft) / aero.nm # [Nm / s]
    # tas = aero.vmach2tas(speed, FL * 100 * aero.ft) / aero.nm # [Nm / s]
    return distance / tas
    # print('tas is: ', tas)
    # print('coefficient is: ', 0.12390501319261213720316622691293)
    # print('coefficient is: ', 0.12377203290246768507638072855464)
    # print('coefficient is: ', 0.12417203290246768507638072855464)
    # time_req = distance / 0.12400203290246768507638072855464
    # return time_req

def time_required2(distances_nm, FL, speed):
    # def _eta2tw_cas_wfl(self, distances_nm, flightlevels_m, cas_m_s):
    """
    Estimate for the given CAS the ETA to the next TW waypoint
    No wind is taken into account.
    :param distances: distances between waypoints
    :param flightlevels: flightlevels for sections
    :param cas_m_s: CAS in m/s
    :return: ETA in seconds
    """
    # cas_m_s = speed * aero.nm / 3600
    # print(cas_m_s)
    distances_m = distances_nm * aero.nm
    # times_s = np.empty_like(distances_m)
    # previous_fl_m = flightlevels_m

    # Assumption is instantanuous climb to next flight level, and instantanuous speed change at new flight level
    # for i, distance_m in enumerate(distances_m):
    #     if flightlevels_m[i] < 0:
    #         next_fl_m = previous_fl_m
    #     else:
    #         next_fl_m = flightlevels_m[i]
    #         previous_fl_m = next_fl_m
    # print('FL is: ', FL)
    next_tas_m_s = aero.vmach2tas(speed, FL * aero.ft * 100)
    # print(next_tas_m_s)
    # next_tas_m_s = aero.vcas2tas(cas_m_s, FL * aero.ft * 100)
    step_time_s = distances_m / (next_tas_m_s + 0.00001)
    # times_s[i] = step_time_s
    # print('step time is: ', step_time_s)
    # total_time_s = np.sum(times_s)
    return step_time_s

'''                 ######################################## 
                    ########################################
                    ###     Upload to Drive Method!!!    ###
                    ########################################   
                    ########################################             
'''
def upload_file(file_upload, name, dir=None):
    gauth = GoogleAuth()
    # Try to load saved client credentials
    gauth.LoadCredentialsFile("mycreds.txt")
    if gauth.credentials is None:
        # Authenticate if they're not there
        gauth.LocalWebserverAuth()
        gauth.SaveCredentialsFile("mycreds.txt")
    elif gauth.access_token_expired:
        # Refresh them if expired
        gauth.Refresh()
    else:
        # Initialize the saved creds
        gauth.Authorize()

    drive = GoogleDrive(gauth)
    if 'min' in dir:        fid = drive_folder_min
    elif 'det' in dir:      fid = drive_folder_det
    elif 'prob' in dir:     fid = drive_folder_prob
    elif 'inf' in dir:      fid = drive_folder_inf
    elif 'input' in dir:    fid = drive_folder_input
    elif 'output' in dir:   fid = drive_folder_output
    else:                   fid = drive_folder_gen
        # print('Upload directory not found!')
        # return

    with open(file_upload, "r") as file:
        file_drive = drive.CreateFile({'parents': [{"kind": "drive#fileLink",
                                                     "id": fid}],
                                        'title': name})
        # os.path.basename(file.name)
        # file_drive.SetContentString(file.read())
        file_drive.SetContentFile(file_upload)
        file_drive.Upload()
        # f = drive.CreateFile({"parents": [{"kind": "drive#fileLink", "id": fid}]})
        # f.SetContentFile(some_path)
        # f.Upload()

        # file1 = drive.CreateFile({'parent': '/home/pi'})
        # file1.SetContentFile('test.txt')
        # file1.Upload()

        #Get content from Google Drive
        # file_list = drive.ListFile({'q': "'< folder >' in parents and trashed=false"}).GetList()
        # file_list = drive.ListFile({'q': "'root' in parents and trashed=false"}).GetList()
        # for file1 in file_list:
        #     print('title: %s, id: %s' % (file1['title'], file1['id']))

    print("File '{}' has been uploaded!".format(file_upload))

'''                 ######################################## 
                    ########################################
                    ###        Excel Writer Method!!!    ###
                    ########################################   
                    ########################################             
'''
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

'''                 ######################################## 
                    ########################################
                    ### Analysis tools for the results!! ###
                    ########################################   
                    ########################################             
'''
# This method reads in the mylog files and orders them on fuelflow for comparison
def compare_ff(file=None):
    if file is None:
        files = os.listdir('output\\')
        files = ['output\\' + files[files.index(i)] for i in files if 'MYLOG' in i]
    else:
        files = list([file])
        files.append('dummy')
    # print(files)
    for i in files:
        if i[-3:] != 'log':
            continue
        # set_of_ac = list()

        with open(i, 'r') as fin:
            data = fin.read().splitlines(True)
            # print(data[0])
        if 'MYLOG' in data[0]:
            with open(i, 'w') as fout:
                fout.writelines(data[1:])
        with open(i, 'r') as f:
            f = pd.read_csv(i)
        # print(f.columns)
        # f = f[2:]
        # print(f)
        f = f.sort_values([' id', '# simt']).reset_index(drop=True)
        # f.reset_index
        set_of_ac = set(f[' id'])
        print('Number of rows in the log: ', len(f))
        # print('The following aircrafts have been found: ', set_of_ac)
        partial_0 = 0
        to_save = i[:-4] + ' - formatted.xlsx'
        writer = pd.ExcelWriter(to_save, engine='xlsxwriter')
        for index, e in enumerate(set_of_ac):
             #int(len(f)/2) +20
            partial_1 = max(partial_0, f[f[' id'] == e].index[0])
            if partial_0 == partial_1:
                partial_1 = -1
            f1 = f[partial_0:partial_1]
            if partial_0 > 0:
                f1 = f1.drop(columns='# simt')
                partial_0 = partial_1
            # f3 = pd.concat([f1[-250:], f2[-250:]], axis=1, ignore_index=True)

            # f1.to_excel('output\\' + i[:-4] + '1.xlsx'.format(i))
            # f2.to_excel('output\\' + i[:-4] + '2.xlsx'.format(i))
            # f1[-100:].to_excel(writer)
            # l = bool(index > 0)
            # print(index)
            # print(len(f.columns))
            f1[-10:].to_excel(writer, startcol=((len(f.columns)+1)*index))
        writer.save()
        # os.startfile(to_save)
    # exit()

# This method creates an overall analysis over all the ensembles
def overall_aggregate(path=None, upload=False):
    if path is None:
        path = os.path.join(os.getcwd(), dest_output)
    with open(os.getcwd() + '/scenario/number_of_ac.txt', "r") as fin:
        number_ac = int(fin.read())
    # path = "C:\Documents\Git 2\output\\runs ----"
    Dir = os.listdir(path)
    to_save = pd.DataFrame()
    skip_list = ['py', 'skip', 'xlogs', 'xls', 'anal']
    spacer = len(set_of_delays) - 1
    to_skip = False
    print(' ')
    for i, dir in enumerate(Dir):
        for skip in skip_list:
            if skip in dir:
                print('skipped: ', dir)
                to_skip = True
                continue
        if to_skip:
            to_skip = False
            continue

        Files = os.listdir(os.path.join(path, dir))
        for m, n in enumerate(Files):
            Files[m] = os.path.join(path, dir, n)

        Files_input_log = os.listdir(os.path.join(path, os.path.split(dest_dir_input_logs)[1], dir))
        for m, n in enumerate(Files_input_log):
            Files_input_log[m] = os.path.join(path, os.path.split(dest_dir_input_logs)[1], dir, n)
        l = 1
        print('Assessing folder {}!'.format(dir))
        for j, file in enumerate(Files):
            if '~$' in file:
                 continue

            if j == 0:
                apple = pd.read_excel(file, index_col=None, header=None)
                banana = list(range(0, len(apple.columns)))
                to_pop = list([0, 1, 2, 3, 5, 6, 7])
                to_pop.reverse()
                for k in to_pop:
                    banana.pop(k)
                apple = apple.drop(columns=banana, axis=1)

                if i == 0:
                    to_save = apple.drop(columns=list([0, 1, 6, 7]))
                    pd_files = pd.read_excel(file, index_col=None, header=None)
                    pd_files_names = list(set([r.pop(0) for r in pd_files[3].astype(str).str.split('-')]))
                    pd_files_names.sort()
                    appleMax = pd.DataFrame(np.ones(number_ac)*0)
                    appleMin = pd.DataFrame(np.ones(number_ac)*1e6)
                else:
                    to_save = pd.concat([to_save, apple[[2, 3, 5]]], axis=1)

                apple = apple.drop(columns=list([0, 1, 2, 3, 5]))
                continue

            apple2 = pd.read_excel(file, index_col=None, header=None)
            apple[6] = pd.to_timedelta(apple[6], unit='s') + pd.to_timedelta(apple2[6], unit='s')
            appleMin = appleMin.clip_upper(apple2[7], axis=0)
            appleMax = appleMax.clip_lower(apple2[7], axis=0)
            apple[7] = apple[7] + apple2[7]
            l += 1

        apple[6] = (apple[6] / l).dt.round('1s')
        apple[7] = round(apple[7] / l, 2)

        # Calculate the standard deviation of the Fuel Consumption
        for j, file in enumerate(Files):
            if '~$' in file:
                continue

            if j == 0:
                apple_0 = pd.read_excel(file, index_col=None, header=None)
                banana = list(range(0, len(apple_0.columns)))
                to_pop = list([0, 1, 2, 3, 5, 6, 7])
                to_pop.reverse()
                for k in to_pop:
                    banana.pop(k)
                apple_0 = apple_0.drop(columns=banana, axis=1)

                if i == 0:
                    pd_files = pd.read_excel(file, index_col=None, header=None)
                    pd_files_names = list(set([r.pop(0) for r in pd_files[3].astype(str).str.split('-')]))
                    pd_files_names.sort()
                    appleStdDev = apple[7]*0
                continue

            apple_1 = pd.read_excel(file, index_col=None, header=None)
            appleStdDev = appleStdDev.add((apple_1[7] - apple[7])**2)

        appleStdDev = (appleStdDev / l)**(1/2)
        pd_min = list()
        pd_max = list()
        for m, n in enumerate(pd_files_names):
            k = n + '-' + dir[2:] + '-0'
            log2 = ''.join(list([line for line in open(Files_input_log[0], 'r') if k in line]))
            logType = ''.join(list([line for line in open(Files_input_log[0], 'r') if 'CRE ' in line]))

            if i == 0:
                appleType = list()
                for q in pd_files_names:
                    apple_0 = logType.find('CRE {}-MIN-0 '.format(q))
                    appleType.append(logType[apple_0:apple_0 + 50].split()[2])

            apple_0 = log2.find('TW_SIZE_AT {}'.format(k))
            apple_1 = log2[apple_0-45:apple_0-10].split()[1]
            apple_2 = int(log2[apple_0:apple_0+50].split()[2])
            apple_3 = datetime.datetime(100, 1, 1, int(apple_1[-8:-6]), int(apple_1[-5:-3]), int(apple_1[-2:]))
            apple_4 = apple_3 - datetime.timedelta(seconds=apple_2 / 2)
            apple_5 = apple_3 + datetime.timedelta(seconds=apple_2 / 2)
            pd_min.append(apple_4.strftime('%H:%M:%S'))
            pd_max.append(apple_5.strftime('%H:%M:%S'))

        pd_min = pd.DataFrame([item for item in pd_min for _ in range(len(set_of_delays))])
        pd_max = pd.DataFrame([item for item in pd_max for _ in range(len(set_of_delays))])
        apple2 = pd.concat([appleMin.round(2), appleMax.round(2), appleStdDev.round(2)], axis=1)
        to_save = pd.concat([to_save, pd_min, pd_max, apple], axis=1)
        del apple, apple2

    print('Saving the results!')
    filename0 = 'meta-analysis.xlsx'#.format(22)
    filename = path + '\\' + 'meta-analysis.xlsx' #.format(22)
    to_save.columns = (['Delay 1', 'Min', ' Dep 1', 'Arr 1 min', 'Arr 1 max', 'Arrival 1', 'Fuel Con 1',
                        'Fuel Min 1', 'Fuel Max 1', 'Fuel StdD 1',
                        'Delay 2', 'Det', ' Dep 2', 'Arr 2 min', 'Arr 2 max', 'Arrival 2', 'Fuel Con 2',
                        'Fuel Min 2', 'Fuel Max 2', 'Fuel StdD 2',
                        'Delay 3', 'Prob', ' Dep 3', 'Arr 3 min', 'Arr 3 max', 'Arrival 3', 'Fuel Con 3',
                        'Fuel Min 3', 'Fuel Max 3', 'Fuel StdD 3',
                        'Delay 4', 'Inf', ' Dep 4', 'Arr 4 min', 'Arr 4 max', 'Arrival 4', 'Fuel Con 4',
                        'Fuel Min 4', 'Fuel Max 4', 'Fuel StdD 4'])
    for dir in Dir:
        if 'min' in dir:    apple = ['Delay 1', 'Min', 'Arrival 1', 'Fuel Con 1']
        elif 'det' in dir:  apple = ['Delay 2', 'Det', 'Arrival 2', 'Fuel Con 2']
        elif 'prob' in dir: apple = ['Delay 3', 'Prob', 'Arrival 3', 'Fuel Con 3']
        elif 'inf' in dir:  apple = ['Delay 4', 'Inf', 'Arrival 4', 'Fuel Con 4']
        else: continue

        dir2 = apple[1]
        to_save[apple[0]] = pd.to_numeric(to_save[apple[0]], errors='coerce')
        to_save[dir2] = [r.pop(0) + '-' + r.pop(0) for r in to_save[dir2].astype(str).str.split('-')]
        to_save2 = to_save[apple]
        to_save2 = to_save2.sort_values(by=[dir2, apple[0]], ascending=True)
        to_save[apple] = to_save2.values

    spacer += 1

    with open(filename, 'wb') as f:
        to_save.index.name = appleType[0]
        to_save[0:spacer].to_excel(f, sheet_name='meta-analysis')

    for k, i in enumerate(range(1, int(number_ac/len(set_of_delays)))):
        j = i * spacer
        l = k + j + 1
        to_save.index.name = appleType[k+1]
        append_df_to_excel(filename, to_save[j:j + spacer], 'meta-analysis', l)

    if upload:
        upload_file(filename, filename0)
    pass

# This method creates an overall analysis over all the ensembles
def overall_aggregate2(path=None, upload=False):
    if path is None:
        path = os.path.join(os.getcwd(), dest_output)
    with open(os.getcwd() + '/scenario/number_of_ac.txt', "r") as fin:
        number_ac = int(fin.read())
    # path = "C:\Documents\Git 2\output\\runs ----"
    Dir = os.listdir(path)
    to_save = pd.DataFrame()
    skip_list = ['py', 'skip', 'xlogs', 'xls', 'anal']
    to_skip = False
    print('\n' + bcolors.HEADER + 'Creating meta-analysis for:\n' + os.path.split(path)[-1] + bcolors.ENDC, '\n')
    for i, dir in enumerate(Dir):
        for skip in skip_list:
            if skip in dir:
                print('skipped: ', dir)
                to_skip = True
                continue
        if to_skip:
            to_skip = False
            continue

        Files = os.listdir(os.path.join(path, dir))
        for m, n in enumerate(Files):
            Files[m] = os.path.join(path, dir, n)

        Files_input_log = os.listdir(os.path.join(path, os.path.split(dest_dir_input_logs)[1], dir))
        for m, n in enumerate(Files_input_log):
            Files_input_log[m] = os.path.join(path, os.path.split(dest_dir_input_logs)[1], dir, n)
        l = 1
        print('Assessing folder {}!'.format(dir))
        for j, file in enumerate(Files):
            if '~$' in file:
                 continue

            if j == 0:
                apple = pd.read_excel(file, index_col=None, header=None)
                if np.isnan(apple[0].iloc[0]):
                    apple = apple.drop([0]).reset_index(drop=True)
                    spacer = len(set(apple[2])) - 1
                else:
                    spacer = len(set(apple[2]))
                banana = list(range(0, len(apple.columns)))
                to_pop = list([0, 1, 2, 3, 5, 6, 7])
                to_pop.reverse()
                for k in to_pop:
                    banana.pop(k)
                apple = apple.drop(columns=banana, axis=1)

                if i == 0:
                    to_save = apple.drop(columns=list([0, 1, 6, 7]))
                    pd_files = pd.read_excel(file, index_col=None, header=None)
                    pd_files = pd_files.drop([0]).reset_index(drop=True)
                    pd_files_names = list(set([r.pop(0) for r in pd_files[3].astype(str).str.split('-')]))
                    pd_files_names.sort()
                    appleMax  = pd.DataFrame(np.ones(number_ac)*0)
                    appleMin  = pd.DataFrame(np.ones(number_ac)*1e6)
                    appleMax2  = pd.DataFrame(np.ones(number_ac)*0)
                    appleMin2  = pd.DataFrame(np.ones(number_ac)*1e8)
                else:
                    to_save = pd.concat([to_save, apple[[2, 3, 5]]], axis=1)

                apple = apple.drop(columns=list([0, 1, 2, 3, 5]))
                continue

            apple2 = pd.read_excel(file, index_col=None, header=None)
            if np.isnan(apple2[0].iloc[0]):
                apple2 = apple2.drop([0]).reset_index(drop=True)
            apple[6] = pd.to_timedelta(apple[6], unit='s') + pd.to_timedelta(apple2[6], unit='s')
            appleMin = appleMin.clip_upper(apple2[7].astype(float), axis=0)
            appleMax = appleMax.clip_lower(apple2[7].astype(float), axis=0)
            appleMin2 = appleMin2.clip_upper(pd.to_timedelta(apple2[6], unit='s').dt.total_seconds(), axis=0)
            appleMax2 = appleMax2.clip_lower(pd.to_timedelta(apple2[6], unit='s').dt.total_seconds(), axis=0)
            apple[7] = apple[7].astype(float) + apple2[7].astype(float)
            l += 1

        apple[6] = (apple[6] / l).dt.round('1s')
        apple[7] = round(apple[7].astype(float) / l, 2)

        # Calculate the standard deviation of the Fuel Consumption and Arrival Time
        for j, file in enumerate(Files):
            if '~$' in file:
                continue

            if j == 0:
                apple_0 = pd.read_excel(file, index_col=None, header=None)
                banana = list(range(0, len(apple_0.columns)))
                to_pop = list([0, 1, 2, 3, 5, 6, 7])
                to_pop.reverse()
                for k in to_pop:
                    banana.pop(k)
                apple_0 = apple_0.drop(columns=banana, axis=1)

                if i == 0:
                    pd_files = pd.read_excel(file, index_col=None, header=None)
                    pd_files = pd_files.drop([0]).reset_index(drop=True)
                    pd_files_names = list(set([r.pop(0) for r in pd_files[3].astype(str).str.split('-')]))
                    pd_files_names.sort()
                    appleStdDev     = apple[7]*0
                    appleStdDev2    = apple[7]*0
                continue

            apple_1 = pd.read_excel(file, index_col=None, header=None)
            apple_1 = apple_1.drop([0]).reset_index(drop=True)
            apple_1[7] = apple_1[7].astype(float)
            appleStdDev     = appleStdDev.add((apple_1[7] - apple[7])**2)
            appleStdDev2    = appleStdDev2.add(((pd.to_timedelta(apple_1[6], unit='s').dt.total_seconds() -
                                                    pd.to_timedelta(apple[6], unit='s').dt.total_seconds()).astype(int))**2)

        appleStdDev     = (appleStdDev / l)**(1/2)
        appleStdDev2    = (appleStdDev2 / l)**(1/2)
        print('Number of Ensembles:', l)
        pd_min  = list()
        pd_max  = list()
        for m, n in enumerate(pd_files_names):
            k = n + '-' + dir[2:] + '-0'
            log2 = ''.join(list([line for line in open(Files_input_log[0], 'r') if k in line]))
            logType = ''.join(list([line for line in open(Files_input_log[0], 'r') if 'CRE ' in line]))

            if i == 0 and m == 0:
                appleType = list()
                for q in pd_files_names:
                    apple_0 = logType.find('CRE {}-MIN-0 '.format(q))
                    appleType.append(logType[apple_0:apple_0 + 50].split()[2])

            apple_0 = log2.find('TW_SIZE_AT {}'.format(k))
            apple_1 = log2[apple_0-45:apple_0-10].split()[1]
            apple_2 = int(log2[apple_0:apple_0+50].split()[2])
            apple_3 = datetime.datetime(100, 1, 1, int(apple_1[-8:-6]), int(apple_1[-5:-3]), int(apple_1[-2:]))
            apple_4 = apple_3 - datetime.timedelta(seconds=apple_2 / 2)
            apple_5 = apple_3 + datetime.timedelta(seconds=apple_2 / 2)
            pd_min.append(apple_4.strftime('%H:%M:%S'))
            pd_max.append(apple_5.strftime('%H:%M:%S'))

        appleMin3 = list()
        appleMax3 = list()
        pd_min = pd.DataFrame([item for item in pd_min for _ in range(len(set_of_delays))])
        pd_max = pd.DataFrame([item for item in pd_max for _ in range(len(set_of_delays))])

        for m in range(len(appleMin2)):
            holder = datetime.datetime(100, 1, 1, 0, 0, 0) + datetime.timedelta(seconds=int(appleMin2.values[m]))
            appleMin3.append(holder.strftime('%H:%M:%S'))
            holder = datetime.datetime(100, 1, 1, 0, 0, 0) + datetime.timedelta(seconds=int(appleMax2.values[m]))
            appleMax3.append(holder.strftime('%H:%M:%S'))

        appleMin3 = pd.DataFrame(appleMin3)
        appleMax3 = pd.DataFrame(appleMax3)
        apple2 = pd.concat([appleMin.round(2), appleMax.round(2), appleStdDev.round(2)], axis=1)
        apple3 = pd.concat([appleMin3, appleMax3, appleStdDev2.round(2)], axis=1)
        to_save = pd.concat([to_save, pd_min, pd_max, apple, apple2, apple3], axis=1)
        del apple, apple2

    filename0 = 'meta-analysis.xlsx'#.format(22)
    filename = path + '\\' + 'meta-analysis-Sep.xlsx' #.format(22)
    print(bcolors.OKGREEN + '\nSaving the results in ' + os.path.split(filename)[-1] + bcolors.ENDC)

    if '3 prob' in Dir:
        to_save.columns = (['Delay 1', 'Min', ' Dep 1', 'Arr 1 min', 'Arr 1 max', 'Arrival 1', 'Fuel Con 1',
                            'Fuel Min 1', 'Fuel Max 1', 'Fuel StdD 1', 'Arr Min 1 ', 'Arr Max 1', 'Arr StdD 1',
                            'Delay 2', 'Det', ' Dep 2', 'Arr 2 min', 'Arr 2 max', 'Arrival 2', 'Fuel Con 2',
                            'Fuel Min 2', 'Fuel Max 2', 'Fuel StdD 2', 'Arr Min 2 ', 'Arr Max 2', 'Arr StdD 2',
                            'Delay 3', 'Prob', ' Dep 3', 'Arr 3 min', 'Arr 3 max', 'Arrival 3', 'Fuel Con 3',
                            'Fuel Min 3', 'Fuel Max 3', 'Fuel StdD 3', 'Arr Min 3 ', 'Arr Max 3', 'Arr StdD 3',
                            'Delay 4', 'Inf', ' Dep 4', 'Arr 4 min', 'Arr 4 max', 'Arrival 4', 'Fuel Con 4',
                            'Fuel Min 4', 'Fuel Max 4', 'Fuel StdD 4', 'Arr Min 4 ', 'Arr Max 4', 'Arr StdD 4'])
    elif '1 min' not in Dir:
        to_save.columns = (['Delay 1', 'Inf', ' Dep 1', 'Arr 1 min', 'Arr 1 max', 'Arrival 1', 'Fuel Con 1',
                            'Fuel Min 1', 'Fuel Max 1', 'Fuel StdD 1', 'Arr Min 1 ', 'Arr Max 1', 'Arr StdD 1'])
    else:
        to_save.columns = (['Delay 1', 'Min', ' Dep 1', 'Arr 1 min', 'Arr 1 max', 'Arrival 1', 'Fuel Con 1',
                            'Fuel Min 1', 'Fuel Max 1', 'Fuel StdD 1', 'Arr Min 1 ', 'Arr Max 1', 'Arr StdD 1',
                            'Delay 2', 'Det', ' Dep 2', 'Arr 2 min', 'Arr 2 max', 'Arrival 2', 'Fuel Con 2',
                            'Fuel Min 2', 'Fuel Max 2', 'Fuel StdD 2', 'Arr Min 2 ', 'Arr Max 2', 'Arr StdD 2',
                            'Delay 3', 'Inf', ' Dep 3', 'Arr 3 min', 'Arr 3 max', 'Arrival 3', 'Fuel Con 3',
                            'Fuel Min 3', 'Fuel Max 3', 'Fuel StdD 3', 'Arr Min 3 ', 'Arr Max 3', 'Arr StdD 3'])

    # Select the columns which have to be ordered together
    for dir in Dir:
        if 'min' in dir:        apple = to_save.columns[00:13] # ['Delay 1', 'Min', 'Arrival 1', 'Fuel Con 1']
        elif 'det' in dir:      apple = to_save.columns[13:26] # ['Delay 2', 'Det', 'Arrival 2', 'Fuel Con 2']
        elif 'prob' in dir:     apple = to_save.columns[26:39] # ['Delay 3', 'Prob', 'Arrival 3', 'Fuel Con 3']
        elif 'inf' in dir:
            if '3 prob' in Dir:
                apple = to_save.columns[39:52] # apple = ['Delay 4', 'Inf', 'Dep 4', 'Arrival 4', 'Fuel Con 4']
            elif '1 min' not in Dir:
                apple = to_save.columns[00:13]
            else:
                apple = to_save.columns[26:39]  # apple = ['Delay 3', 'Inf', 'Arrival 3', 'Fuel Con 3']
        else: continue

        dir2 = apple[1]
        to_save[apple[0]] = pd.to_numeric(to_save[apple[0]], errors='coerce')
        to_save[dir2] = [r.pop(0) + '-' + r.pop(0) for r in to_save[dir2].astype(str).str.split('-')]
        to_save2 = to_save[apple]
        to_save2 = to_save2.sort_values(by=[dir2, apple[0]], ascending=True)
        to_save[apple] = to_save2.values
    spacer += 1

    with open(filename, 'wb') as f:
        to_save.index.name = appleType[0]
        to_save[0:spacer].to_excel(f, sheet_name='meta-analysis')

    for k, i in enumerate(range(1, int(number_ac/spacer))):
        j = i * spacer
        l = k + j + 1
        to_save.index.name = appleType[k+1]
        append_df_to_excel(filename, to_save[j:j + spacer], 'meta-analysis', l)

    if upload:
        upload_file(filename, filename0)
    pass

# This method creates an analysis per ensemble
def result_analysis(path=None, skip_dir=False, upload=False, skip_flights='zero'):

    if path is None:
        path = os.path.join(os.getcwd(), dest_output)

    path_analysis = os.path.join(path, 'analysis')
    if not os.path.isdir(path_analysis):
        os.makedirs(path_analysis)

    Dir = os.listdir(path)
    to_save = pd.DataFrame()
    if skip_dir:
        skip_list = ['py', 'skip', 'put', 'xls', 'anal'] + list(skip_dir)
    else:
        skip_list = ['py', 'skip', 'put', 'xls', 'anal']
    to_skip = False
    cruise_speed = pd.read_excel(cruise_file)

    for i, dir in enumerate(Dir):
        for skip in skip_list:
            if skip in dir:
                to_skip = True
                continue
        if to_skip:
            print('skipped: ', dir)
            to_skip = False
            continue

        Files = os.listdir(os.path.join(path, dir))
        for m, n in enumerate(Files):
            Files[m] = os.path.join(path, dir, n)

        if path:
            Files_input_log = os.listdir(os.path.join(path, os.path.split(dest_dir_input_logs)[1], dir))
            for m, n in enumerate(Files_input_log):
                Files_input_log[m] = os.path.join(path, os.path.split(dest_dir_input_logs)[1], dir, n)
        else:
            Files_input_log = os.listdir(os.path.join(os.getcwd(), dest_dir_input_logs, dir))
            for m, n in enumerate(Files_input_log):
                Files_input_log[m] = os.path.join(os.getcwd(), dest_dir_input_logs, dir, n)
        # print('input log: ', Files_input_log)

        pd_files = pd.read_excel(Files[0], index_col=None, header=None)
        pd_files_names = list(set([r.pop(0) for r in pd_files[3].astype(str).str.split('-')]))
        pd_files_names.sort()
        legend = (['', '',
                   'colour blue if RTA',
                   'colour red if too late for TW',
                   'colour yellow if too early for TW',
                   'colour green if within TW', '', '', '', ''])

        for k in pd_files_names:
            if k in skip_flights:
                continue
            print('\nStarting trajectory {} of directory {}!'.format(k, dir))
            filename0 = '{}{}.xlsx'.format(k, dir[1:])
            filename = os.path.join(path_analysis, filename0)
            to_save = to_save.reindex(to_save.columns.tolist() + ['Name', 'Vstart'])

            # Read in both files
            # print(Files, Files_input_log)
            for a, b in zip(range(0, len(Files_input_log), 10), range(10, len(Files_input_log)+1, 10)):
                print(' ')
                print('Start = ', a, '| Stop = ', b)
                filename0 = '{}{} - {}.xlsx'.format(k, dir[1:], b)
                filename = os.path.join(path_analysis, filename0)
                for m, n in zip(Files[a:b], Files_input_log[a:b]):
                    ExcelDoc = pd.read_excel(m, index_col=None, header=None)
                    Ensemble = ExcelDoc[1][0]
                    Flights = ExcelDoc[ExcelDoc[3].str.contains(k)].reset_index(drop=True).sort_values(2)\
                        .dropna(axis='columns').drop(columns=list([0, 1, 4, 8]))
                    Flights = Flights.rename(index=str, columns={
                        2: "Delay", 3: "Name", 5: "Start Time", 6: "End Time", 7: "Fuel Used [kg]"})
                    log2 = ''.join(list([line for line in open(n, 'r') if k in line]))

                    # Add Columns
                    holder = list()
                    FL_start = list()
                    for o, p in enumerate(Flights['Name']):
                        apple = log2.find('CRE ' + p + ' ')
                        holder.append(log2[apple:apple + 70].split()[5])
                        FL_start.append(log2[apple:apple + 100].split()[6][:3])
                    Flights['Vstart [Mach]'] = None
                    actype = log2[apple:apple + 100].split()[2]
                    for o, _ in enumerate(Flights['Vstart [Mach]']):
                        # Flights['Vstart [Mach]'][o] = round(aero.vcas2mach(int(holder[o]) * aero.nm / 3600,
                        #                                                    int(FL_start[o]) * 100 * aero.ft), 2)
                        Flights['Vstart [Mach]'][o] = \
                            cruise_speed[cruise_speed['AC Type'] == actype]['FL' + FL_start[o]].item()
                    cols = Flights.columns.tolist()
                    cols.insert(2, cols.pop(-1))
                    Flights = Flights[cols]

                    # Rename Columns
                    FL = list()
                    pd_min = list()
                    pd_max = list()
                    for o, p in enumerate(range(len(cols)-6)):
                        if o == 0:
                            apple = log2.find('CRE {}-{}-0'.format(k, dir[2:].upper()))
                            actype = log2[apple:apple + 100].split()[2]
                            FL.append('FL' + log2[apple:apple + 100].split()[6][:3])
                            apple = log2.find('RTA_AT {}-{}-0'.format(k, dir[2:]))
                            apple = log2[apple:apple+50].split()[3][:8]
                            apple = datetime.datetime(100, 1, 1, int(apple[-8:-6]), int(apple[-5:-3]), int(apple[-2:]))
                            banana = log2.find('TW_SIZE_AT {}-{}-0'.format(k, dir[2:]))
                            banana = int(log2[banana:banana+50].split()[2])
                            if TW_place:
                                apple1 = apple - datetime.timedelta(seconds=banana/2)
                                apple2 = apple + datetime.timedelta(seconds=banana/2)
                            else:
                                apple1 = apple #- datetime.timedelta(seconds=banana/2)
                                apple2 = apple + datetime.timedelta(seconds=banana)
                            pd_min.append(apple1.strftime('%H:%M:%S'))
                            pd_max.append(apple2.strftime('%H:%M:%S'))
                        else:
                            apple = log2.find('ADDWPT {}-{}-0-{}'.format(k, dir[2:], o))
                            FL.append(log2[apple:apple + 100].split()[2])
                            apple = log2.find('RTA_AT {}-{}-0-{}'.format(k, dir[2:], o))
                            apple = log2[apple:apple+50].split()[2]
                            banana = log2.find('TW_SIZE_AT {}-{}-0-{}'.format(k, dir[2:], o))
                            # print(log2[banana:banana + 50].split())
                            banana = int(log2[banana:banana + 50].split()[2])
                            # print('banana is: ', banana)
                            apple = datetime.datetime(100, 1, 1, int(apple[-8:-6]), int(apple[-5:-3]), int(apple[-2:]))
                            apple1 = apple - datetime.timedelta(seconds=banana/2)
                            apple2 = apple + datetime.timedelta(seconds=banana/2)
                            pd_min.append(apple1.strftime('%H:%M:%S'))
                            pd_max.append(apple2.strftime('%H:%M:%S'))

                    for o, p in enumerate(range(len(FL))):
                        q = f'WP{o} [{FL[o]}]' #.format(str(o), FL[o])
                        Flights = Flights.rename(index=str, columns={(o+9): q})

                    # Add Rows
                    holder1 = list(['-', 'Goal Early', '-', pd_min[0], pd_min[-1], '-'])
                    holder1.extend(pd_min)
                    holder2 = ['-', 'Goal Late', '-', pd_max[0], pd_max[-1], '-']
                    holder2.extend(pd_max)
                    Flights.ix[len(Flights):len(Flights)+2] = np.nan
                    Flights = Flights.append(pd.Series(), ignore_index=True)
                    Flights = Flights.append(pd.Series(), ignore_index=True)
                    Flights = Flights.shift(2)
                    Flights.ix[0] = pd.Series(holder1, index=Flights.columns)
                    Flights.ix[1] = pd.Series(holder2, index=Flights.columns)

                    # Add Column with FT
                    holder = list()
                    for o, p in zip(Flights['Start Time'], Flights['End Time']):
                        apple = datetime.datetime(100, 1, 1, int(o[-8:-6]), int(o[-5:-3]), int(o[-2:]))
                        banana = datetime.datetime(100, 1, 1, int(p[-8:-6]), int(p[-5:-3]), int(p[-2:]))
                        citrus = banana - apple
                        hours = divmod(int(citrus.total_seconds()), 3600)  # Use remainder of days to calc hours
                        minutes = divmod(hours[1], 60)  # Use remainder of hours to calc minutes
                        seconds = divmod(minutes[1], 1)
                        holder.append('{}:{}:{}'.format(hours[0], minutes[0], seconds[0]))
                    holder[0] = actype
                    Flights['nWP = {}'.format(Flights.shape[1]-6)] = holder
                    Flights['Legend'] = legend

                    # Create second Waypoint Analysis with Normalised Time
                    Flights2 = Flights.copy()
                    for q, r in enumerate(Flights2):
                        if 'FL' not in r:
                            continue
                        holder = list()
                        for o, p in zip(Flights2['Start Time'], Flights2[r]):
                            apple = datetime.datetime(100, 1, 1, int(o[-8:-6]), int(o[-5:-3]), int(o[-2:]))
                            banana = datetime.datetime(100, 1, 1, int(p[-8:-6]), int(p[-5:-3]), int(p[-2:]))
                            citrus = banana - apple
                            # days = divmod(, 86400)  # Get days (without [0]!)
                            hours = divmod(int(citrus.total_seconds()), 3600)  # Use remainder of days to calc hours
                            minutes = divmod(hours[1], 60)  # Use remainder of hours to calc minutes
                            seconds = divmod(minutes[1], 1)
                            holder.append('{}:{}:{}'.format(hours[0], str(minutes[0]).zfill(2),
                                                            str(seconds[0]).zfill(2)))
                        Flights2[r] = holder

                    # Create Speed Input Analysis
                    Speed_input2 = pd.DataFrame()
                    for s, q in enumerate(Flights['Name'][2:]):
                        obj = [[] for i in range(3)]
                        Speed_input = pd.DataFrame()
                        X = 'SPD2 {},'.format(q)
                        Y = 'SPD {} '.format(q)
                        indices1 = indices(log2, X)
                        indices2 = indices(log2, Y)
                        obj[0].append('# {}s Delay'.format(q.split('-')[2]))
                        obj[0].append('Time')
                        obj[1].append('# {} inputs'.format(len(indices1)))
                        obj[1].append('SPD2')
                        obj[2].append('# V = {}'.format(Flights['Vstart [Mach]'][s+2]))
                        obj[2].append('SPD')

                        for o, (p, r) in enumerate(zip(indices1, indices2)):
                            obj[0].append(log2[p-12:p+30].split()[0][:8])
                            obj[1].append(log2[p - 12:p + 30].split()[2])
                            obj[2].append(log2[r - 12:r + 30].split()[2])

                        Speed_input[0] = obj[0]
                        Speed_input[1] = obj[1]
                        Speed_input[2] = obj[2]
                        Speed_input2 = pd.concat([Speed_input2, Speed_input], ignore_index=True, axis=1)
                    Speed_input = Speed_input2.replace(np.nan, '', regex=True)

                    # Apply color mapping
                    Flights = Flights.style.apply(color, subset=list(Flights.columns[6:-2]))
                    Flights2 = Flights2.style.apply(color, subset=list(Flights.columns[6:-2]))

                    # Write away the Analysis Tables
                    sheet = 'Ensemble=' + str(Ensemble).zfill(2)

                    # Waypoint Analysis
                    append_df_to_excel(filename, pd.DataFrame(['Waypoint Analysis']), sheet,
                                       None, False, header=False, index=False)
                    append_df_to_excel(filename, Flights, sheet)

                    # Waypoint Analysis with Normalised Time
                    append_df_to_excel(filename, pd.DataFrame(['Waypoint Analysis with Normalised Time']), sheet,
                                       None, False, header=False, index=False)
                    append_df_to_excel(filename, Flights2, sheet)

                    # Speed Input Analysis
                    append_df_to_excel(filename, pd.DataFrame(['Speed Input']), sheet,
                                       None, False, header=False, index=False)
                    append_df_to_excel(filename, Speed_input, sheet, None, False, header=False)

                    print('Finished Ensemble {}!'.format(Ensemble))
            print('Finished Trajectory {}!'.format(k))
            # os.startfile(filename)

            if upload:
                upload_file(filename, filename0)
    pass

# This method creates an analysis per ensemble
def result_analysis2(path=None, skip_dir=False, upload=False, skip_flights='zero'):
    if path is None:
        path = os.path.join(os.getcwd(), dest_output)

    path_analysis = os.path.join(path, 'analysis')
    if not os.path.isdir(path_analysis):
        os.makedirs(path_analysis)

    Dir = os.listdir(path)
    to_save = pd.DataFrame()
    if skip_dir:
        skip_list = ['py', 'skip', 'put', 'xls', 'anal'] + list(skip_dir)
    else:
        skip_list = ['py', 'skip', 'put', 'xls', 'anal']
    to_skip = False
    stepsize = 10
    cruise_speed = pd.read_excel(cruise_file2)

    for i, dir in enumerate(Dir):
        for skip in skip_list:
            if skip in dir:
                to_skip = True
                continue
        if to_skip:
            print('skipped: ', dir)
            to_skip = False
            continue

        Files = os.listdir(os.path.join(path, dir))
        for m, n in enumerate(Files):
            Files[m] = os.path.join(path, dir, n)

        if path:
            Files_input_log = os.listdir(os.path.join(path, os.path.split(dest_dir_input_logs)[1], dir))
            for m, n in enumerate(Files_input_log):
                Files_input_log[m] = os.path.join(path, os.path.split(dest_dir_input_logs)[1], dir, n)
        else:
            Files_input_log = os.listdir(os.path.join(os.getcwd(), dest_dir_input_logs, dir))
            for m, n in enumerate(Files_input_log):
                Files_input_log[m] = os.path.join(os.getcwd(), dest_dir_input_logs, dir, n)
        # print('input log: ', Files_input_log)

        pd_files = pd.read_excel(Files[0], index_col=None, header=None)
        pd_files = pd_files.drop([0]).reset_index(drop=True)
        pd_files_names = list(set([r.pop(0) for r in pd_files[3].astype(str).str.split('-')]))
        pd_files_names.sort()
        legend = ([ ' ', ' ',
                    'colour green if within TW', ' ', ' ',
                    'colour yellow if too early for TW', ' ',
                    'colour red if too late for TW',
                    ' '])
        print(' ')
        for kk, k in enumerate(pd_files_names):
            if k in skip_flights:
                continue
            # if kk == 4:
            #     return
            print('Starting trajectory {} of directory {}!'.format(k, dir))
            filename0 = '{}{}.xlsx'.format(k, dir[1:])
            filename = os.path.join(path_analysis, filename0)
            to_save = to_save.reindex(to_save.columns.tolist() + ['Name', 'Vstart'])

            # Read in both files
            for a, b in zip(range(0, len(Files_input_log), stepsize), range(stepsize, len(Files_input_log)+1, stepsize)):
                # if b > 3:
                #     return
                print('\nStart = ', a, '| Stop = ', b)
                filename0 = '{}{} - {}.xlsx'.format(k, dir[1:], b)
                filename = os.path.join(path_analysis, filename0)
                for m, n in zip(Files[a:b], Files_input_log[a:b]):
                    ExcelDoc = pd.read_excel(m, index_col=None, header=None)
                    ExcelDoc = ExcelDoc.drop([0]).reset_index(drop=True)
                    Ensemble = ExcelDoc[1][0]
                    Flights = ExcelDoc[ExcelDoc[3].str.contains(k)].reset_index(drop=True).sort_values(2)\
                        .dropna(axis='columns').drop(columns=list([0, 1, 4, 8]))
                    Flights = Flights.rename(index=str, columns={
                        2: "Delay", 3: "Name", 5: "Start Time", 6: "End Time", 7: "Fuel Used [kg]"})
                    log2 = ''.join(list([line for line in open(n, 'r') if k in line]))

                    # Add Columns
                    holder = list()
                    FL_start = list()
                    for o, p in enumerate(Flights['Name']):
                        apple = log2.find('CRE ' + p + ' ')
                        holder.append(log2[apple:apple + 70].split()[5])
                        FL_start.append(log2[apple:apple + 100].split()[6][:3])
                    Flights['Vstart [Mach]'] = None
                    actype = log2[apple:apple + 100].split()[2]
                    for o, _ in enumerate(Flights['Vstart [Mach]']):
                        Flights['Vstart [Mach]'][o] = cruise_speed[cruise_speed['AC Type'] == actype][
                                'FL' + str(round(int(FL_start[o]), -1))].item()
                    cols = Flights.columns.tolist()
                    cols.insert(2, cols.pop(-1))
                    Flights = Flights[cols]

                    # Rename Columns
                    FL = list()
                    pd_min = list()
                    pd_max = list()
                    for o, p in enumerate(range(len(cols)-6)):
                        if o == 0:
                            apple = log2.find('CRE {}-{}-0'.format(k, dir[2:].upper()))
                            actype = log2[apple:apple + 100].split()[2]
                            FL.append('FL' + log2[apple:apple + 100].split()[6][:3])
                            apple = log2.find('RTA_AT {}-{}-0'.format(k, dir[2:]))
                            apple = log2[apple:apple+50].split()[3][:8]
                            apple = datetime.datetime(100, 1, 1, int(apple[-8:-6]), int(apple[-5:-3]), int(apple[-2:]))
                            banana = log2.find('TW_SIZE_AT {}-{}-0'.format(k, dir[2:]))
                            banana = int(log2[banana:banana+50].split()[2])
                            if TW_place:
                                apple1 = apple - datetime.timedelta(seconds=banana/2)
                                apple2 = apple + datetime.timedelta(seconds=banana/2)
                            else:
                                apple1 = apple #- datetime.timedelta(seconds=banana/2)
                                apple2 = apple + datetime.timedelta(seconds=banana)
                            pd_min.append(apple1.strftime('%H:%M:%S'))
                            pd_max.append(apple2.strftime('%H:%M:%S'))
                        else:
                            apple = log2.find('ADDWPT {}-{}-0-{}'.format(k, dir[2:], o))
                            FL.append(log2[apple:apple + 100].split()[2])
                            apple = log2.find('RTA_AT {}-{}-0-{}'.format(k, dir[2:], o))
                            apple = log2[apple:apple+50].split()[2]
                            banana = log2.find('TW_SIZE_AT {}-{}-0-{}'.format(k, dir[2:], o))
                            # print(log2[banana:banana + 50].split())
                            banana = int(log2[banana:banana + 50].split()[2])
                            # print('banana is: ', banana)
                            apple = datetime.datetime(100, 1, 1, int(apple[-8:-6]), int(apple[-5:-3]), int(apple[-2:]))
                            apple1 = apple - datetime.timedelta(seconds=banana/2)
                            apple2 = apple + datetime.timedelta(seconds=banana/2)
                            pd_min.append(apple1.strftime('%H:%M:%S'))
                            pd_max.append(apple2.strftime('%H:%M:%S'))

                    for o, p in enumerate(range(len(FL))):
                        q = f'WP{o} [{FL[o]}]'
                        Flights = Flights.rename(index=str, columns={(o+9): q})

                    # Add Rows
                    holder1 = list(['-2', 'Goal Early', '-', pd_min[0], pd_min[-1], '-'])
                    holder1.extend(pd_min)
                    holder2 = ['-1', 'Goal Late', '-', pd_max[0], pd_max[-1], '-']
                    holder2.extend(pd_max)
                    Flights.ix[len(Flights):len(Flights)+2] = np.nan
                    Flights = Flights.append(pd.Series(), ignore_index=True)
                    Flights = Flights.append(pd.Series(), ignore_index=True)
                    Flights = Flights.shift(2)
                    Flights.ix[0] = pd.Series(holder1, index=Flights.columns)
                    Flights.ix[1] = pd.Series(holder2, index=Flights.columns)

                    # Add Column with FT
                    holder = list()
                    for o, p in zip(Flights['Start Time'], Flights['End Time']):
                        apple = datetime.datetime(100, 1, 1, int(o[-8:-6]), int(o[-5:-3]), int(o[-2:]))
                        banana = datetime.datetime(100, 1, 1, int(p[-8:-6]), int(p[-5:-3]), int(p[-2:]))
                        citrus = banana - apple
                        hours = divmod(int(citrus.total_seconds()), 3600)  # Use remainder of days to calc hours
                        minutes = divmod(hours[1], 60)  # Use remainder of hours to calc minutes
                        seconds = divmod(minutes[1], 1)
                        holder.append('{}:{}:{}'.format(hours[0], minutes[0], seconds[0]))
                    holder[0] = actype
                    Flights['nWP = {}'.format(Flights.shape[1]-6)] = holder
                    for o in range(Flights.shape[0]+1):
                        if o > len(legend):
                            legend = legend + list(' ')
                    Flights['Legend'] = legend
                    Flights['Delay'] = Flights['Delay'].astype(float)
                    Flights.sort_values(by='Delay', inplace=True)
                    Flights.reset_index(drop=True, inplace=True)

                    # Create second Waypoint Analysis with Normalised Time
                    Flights2 = Flights.copy()
                    for q, r in enumerate(Flights2):
                        if 'FL' not in r:
                            continue
                        holder = list()
                        for o, p in zip(Flights2['Start Time'], Flights2[r]):
                            apple = datetime.datetime(100, 1, 1, int(o[-8:-6]), int(o[-5:-3]), int(o[-2:]))
                            banana = datetime.datetime(100, 1, 1, int(p[-8:-6]), int(p[-5:-3]), int(p[-2:]))
                            citrus = banana - apple
                            # days = divmod(, 86400)  # Get days (without [0]!)
                            hours = divmod(int(citrus.total_seconds()), 3600)  # Use remainder of days to calc hours
                            minutes = divmod(hours[1], 60)  # Use remainder of hours to calc minutes
                            seconds = divmod(minutes[1], 1)
                            holder.append('{}:{}:{}'.format(hours[0], str(minutes[0]).zfill(2),
                                                            str(seconds[0]).zfill(2)))
                        Flights2[r] = holder

                    # Create Speed Input Analysis
                    Speed_input2 = pd.DataFrame()
                    for s, q in enumerate(Flights['Name'][2:]):
                        obj = [[] for i in range(3)]
                        Speed_input = pd.DataFrame()
                        X = 'SPD2 {},'.format(q)
                        Y = 'SPD {} '.format(q)
                        indices1 = indices(log2, X)
                        indices2 = indices(log2, Y)
                        obj[0].append('# {}s Delay'.format(q.split('-')[2]))
                        obj[0].append('Time')
                        obj[1].append('# {} inputs'.format(len(indices1)))
                        obj[1].append('SPD2')
                        obj[2].append('# V = {}'.format(Flights['Vstart [Mach]'][s+2]))
                        obj[2].append('SPD')

                        for o, (p, r) in enumerate(zip(indices1, indices2)):
                            obj[0].append(log2[p-12:p+30].split()[0][:8])
                            obj[1].append(log2[p - 12:p + 30].split()[2])
                            obj[2].append(log2[r - 12:r + 30].split()[2])

                        Speed_input[0] = obj[0]
                        Speed_input[1] = obj[1]
                        Speed_input[2] = obj[2]
                        Speed_input2 = pd.concat([Speed_input2, Speed_input], ignore_index=True, axis=1)
                    Speed_input = Speed_input2.replace(np.nan, '', regex=True)

                    # Apply color mapping
                    Flights = Flights.style.apply(color, subset=list(Flights.columns[6:-2]))
                    Flights2 = Flights2.style.apply(color, subset=list(Flights.columns[6:-2]))

                    # Write away the Analysis Tables
                    sheet = 'Ensemble=' + str(Ensemble).zfill(2)

                    # Waypoint Analysis
                    append_df_to_excel(filename, pd.DataFrame(['Waypoint Analysis']), sheet,
                                       None, False, header=False, index=False)
                    append_df_to_excel(filename, Flights, sheet)

                    # Waypoint Analysis with Normalised Time
                    append_df_to_excel(filename, pd.DataFrame(['Waypoint Analysis with Normalised Time']), sheet,
                                       None, False, header=False, index=False)
                    append_df_to_excel(filename, Flights2, sheet)

                    # Speed Input Analysis
                    append_df_to_excel(filename, pd.DataFrame(['Speed Input']), sheet,
                                       None, False, header=False, index=False)
                    append_df_to_excel(filename, Speed_input, sheet, None, False, header=False)

                    print('Finished Ensemble {}!'.format(Ensemble))
            print('Finished Trajectory {}!\n'.format(k))
            # os.startfile(filename)

            if upload:
                upload_file(filename, filename0)
    pass

# Create a function that provides the filtered logs for a selected trajectory for plot-purposes
# WPonly: True = First Wind measurement at WP only, False = Every Wind measurement
# Choice: True = Magnitude + Direction, False = V north + Veast
def getLog(path=None, traj='AZA1572', selection=[0, 300, 600], wponly=True, ensembles=None, choice=False):
    print('Processing flight {}!'.format(traj))

    # Setup variables
    if path == None:
        path = 'C:\\Documents\\Git 2\\output\\runs\\xlogs input\\4 inf'
    else:
        path = os.path.join(path, 'xlogs input\\1 min')
    Dir = os.listdir(path)
    if ensembles:
        print(f'Plot is limited to the first {ensembles} ensembles!')
        Dir = Dir[0:ensembles]
    file_path = os.path.join(path, Dir[0])
    # log = ''.join(list([line for line in open(file_path, 'r') if 'CRE ' in line]))

    # Find every delayed sibling-trajectory of the selected one
    names = list()
    for line in open(file_path, 'r'):
        for word in line.split():
            if traj in word:
                if len(word.split('-')) < 4 and ':' not in word and ',' not in word:
                    names.append(word)
    names = list(set(i.upper() for i in names))
    names.sort(key=lambda x: x.split("-", 2)[-1].zfill(4))

    # Store the wind-data of every ensemble in a dataframe
    if choice:
        df = pd.DataFrame(columns=['Speed [m/s]', 'Direction [deg]', 'Delay', 'Waypoint'])
    else:
        df = pd.DataFrame(columns=['Speed [m/s]', 'North / East', 'Delay', 'Waypoint'])
    counter = 0
    for ensemble, ensembles in enumerate(Dir):
        counter += 1
        file_path = os.path.join(path, ensembles)
        for name_0, name in enumerate(names):
            log = ''.join(list([line for line in open(file_path, 'r') if 'ECHO {} '.format(name) in line]))
            holder = None
            # div = 1
            for line_0, line in enumerate(log.split('\n')):
                if line_0 % 2 == 0:
                    continue
                elif float(line.split()[2]) != holder:
                    # if line_0 % div == 0:
                        if choice:
                            # Speed and Direction Version
                            speed = np.sqrt(float(line.split(' ')[3][1:-1])**2 + float(line.split(' ')[4][1:-1])**2)
                            direction = np.arctan(float(line.split(' ')[4][1:-1])/float(line.split(' ')[3][1:-1])) * 180/np.pi
                            df = df.append(pd.DataFrame([[speed, direction, int(name.split('-')[2]),
                                                          int(float(line.split(' ')[2]))]], columns=df.columns ), ignore_index=True)
                        else:
                            # V north and V east Version
                            speed = float(line.split(' ')[3][1:-1])
                            direction = str('V north')
                            df = df.append(pd.DataFrame([[speed, direction, int(name.split('-')[2]),
                                                          int(float(line.split(' ')[2]))]], columns=df.columns), ignore_index=True)
                            speed = float(line.split(' ')[4][1:-1])
                            direction = str('V east')
                            df = df.append(pd.DataFrame([[speed, direction, int(name.split('-')[2]),
                                                          int(float(line.split(' ')[2]))]], columns=df.columns), ignore_index=True)
                        if wponly:
                            holder = float(line.split()[2])

    print(f"{counter} ensembles have been processed!")
    if choice:
        # Speed and Direction Plots
        dfsub = df[df['delay'].isin(selection)]
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(20, 10))

        sns.stripplot(x="Waypoint", y="Speed [m/s]", hue="Delay", data=dfsub,
                      jitter=True, dodge=True, color=".3", ax=ax1)
        sns.boxplot(x="Waypoint", y='Speed [m/s]', data=dfsub, hue="Delay",
                    ax=ax1, showfliers=False)
        handles, labels = ax1.get_legend_handles_labels()
        ax1.legend(handles[0:len(selection)], labels[0:len(selection)], title='delay in seconds')
        ax1.set_title('Magnitude of the Wind', fontsize=30)

        sns.stripplot(x="Waypoint", y='Direction [deg]', hue="Delay", data=dfsub,
                      jitter=True, dodge=True, color=".3", ax=ax2)
        sns.boxplot(x="Waypoint", y='Direction [deg]', data=dfsub, hue="Delay",
                    ax=ax2, showfliers=False)
        handles, labels = ax2.get_legend_handles_labels()
        ax2.legend(handles[0:len(selection)], labels[0:len(selection)], title='delay in seconds')
        ax2.set_title('Direction of the Wind', fontsize=30)

        print('Flight {} is shown!\n'.format(traj))
        plt.tight_layout()
        plt.show()
    else:
        # V north and V east Plots
        dot_size = 2
        dfsub = df[df['Delay'].isin(selection)]
        dfsub = dfsub[dfsub['North / East'] == 'V north']
        dfsub = dfsub[(dfsub['Waypoint'] % 3 == 1) | (dfsub['Waypoint'] > 22)] # partial selection of waypoints
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(5, 9))
        # print(dfsub)
        sns.stripplot(x="Waypoint", y="Speed [m/s]", hue="Delay", data=dfsub,
                      jitter=True, dodge=True, color=".3", ax=ax1, size=dot_size)
        sns.boxplot(x="Waypoint", y='Speed [m/s]', data=dfsub, hue="Delay",
                    ax=ax1, showfliers=False)
        handles, labels = ax1.get_legend_handles_labels()
        ax1.legend(handles[0:len(selection)], labels[0:len(selection)], fontsize=8,
                   title='delay [seconds]', loc='upper left', framealpha=0.1)
        ax1.set_title('North-component of the Wind', fontsize=16)
        # ax1.axes.set_title(fontsize=50)
        ax1.set_xlabel("Waypoint [#]", fontsize=0)
        ax1.set_ylabel('Speed [m/s]', fontsize=14)
        ax1.tick_params(labelsize=14)
        # ax1.ytick_params(labelsize=10)

        dfsub = df[df['Delay'].isin(selection)]
        dfsub = dfsub[dfsub['North / East'] == 'V east']
        dfsub = dfsub[(dfsub['Waypoint'] % 3 == 1) | (dfsub['Waypoint'] > 22)] # partial selection of waypoints
        sns.stripplot(x="Waypoint", y="Speed [m/s]", hue="Delay", data=dfsub,
                      jitter=True, dodge=True, color=".3", ax=ax2, size=dot_size)
        sns.boxplot(x="Waypoint", y="Speed [m/s]", data=dfsub, hue="Delay",
                    ax=ax2, showfliers=False)
        handles, labels = ax2.get_legend_handles_labels()
        ax2.legend(handles[0:len(selection)], labels[0:len(selection)], fontsize=8,
                   title='delay [seconds]', loc='bottom left', framealpha=0.1)
        ax2.set_title('East-component of the Wind', fontsize=16)
        ax2.set_xlabel("Waypoint [#]", fontsize=14)
        ax2.set_ylabel('Speed [m/s]', fontsize=14)
        ax2.tick_params(labelsize=14)

        print('Flight {} is shown!\n'.format(traj))
        plt.tight_layout()
        plt.show()

    return

def indices(lst, element):
    result = []
    offset = -1
    while True:
        try:
            offset = lst.index(element, offset + 1)
        except ValueError:
            return result
        result.append(offset)

def color(current):
    early = current[0]
    late = current[1]
    apple = datetime.datetime(100, 1, 1, int(early[-8:-6]), int(early[-5:-3]), int(early[-2:]))
    banana = datetime.datetime(100, 1, 1, int(late[-8:-6]), int(late[-5:-3]), int(late[-2:]))
    colours = list()
    # print(current)
    for j, i in enumerate(current):
        if j == 0 or j == 1:
            colours.append('')
            continue
        citrus = datetime.datetime(100, 1, 1, int(i[-8:-6]), int(i[-5:-3]), int(i[-2:]))
        if citrus > banana:
            colours.append('background-color: brown')
        elif citrus < apple:
            colours.append('background-color: yellow')
        else:
            colours.append('background-color: green')
    # print(colours)
    # NAMED_COLORS = {
    #     'maroon': '800000',
    #     'brown': 'A52A2A',
    #     'red': 'FF0000',
    #     'pink': 'FFC0CB',
    #     'orange': 'FFA500',
    #     'yellow': 'FFFF00',
    #     'olive': '808000',
    #     'green': '008000',
    #     'purple': '800080',
    #     'fuchsia': 'FF00FF',
    #     'lime': '00FF00',
    #     'teal': '008080',
    #     'aqua': '00FFFF',
    #     'blue': '0000FF',
    #     'navy': '000080',
    #     'black': '000000',
    #     'gray': '808080',
    #     'grey': '808080',
    #     'silver': 'C0C0C0',
    #     'white': 'FFFFFF',
    # }
    return colours

def color_min(s):
    apple = datetime.datetime(100, 1, 1, int(s[0][-8:-6]), int(s[0][-5:-3]), int(s[0][-2:])) + datetime.timedelta(seconds=1)
    below_min = s
    for i, j in enumerate(s):
        if i == 0 or i == 1:
            below_min[i] = False
            continue
        banana = datetime.datetime(100, 1, 1, int(s[i][-8:-6]), int(s[i][-5:-3]), int(s[i][-2:])) \
                 + datetime.timedelta(seconds=1)
        if banana.total_seconds() < apple.total_seconds():
            below_min[i] = True
    return ['background-color: yellow' if v else '' for v in below_min]

def autolabel(rects, ax, flag=True, font=8, ratio=1.02):
    """
    Attach a text label above each bar displaying its height
    """
    if flag:
        for rect in rects:
            height = rect.get_height()
            ax.text(rect.get_x() + rect.get_width() / 2., ratio * height,
                    '%d' % int(round(height)),
                    ha='center', va='bottom', fontsize=font)
    else:
        for rect in rects:
            height = rect.get_height()
            if height == 0:
                ax.text(rect.get_x() + rect.get_width() / 2., ratio * height,
                        f'{int(height)}', ha='center', va='bottom', fontsize=font)
            else:
                ax.text(rect.get_x() + rect.get_width() / 2., ratio * height,
                    f'{round(height, 1)}', ha='center', va='bottom', fontsize=font)

# Create a plot with the fuel usages per delay
def fuelvsdelay(flight, path=None, ylimits=None):
    print(f'\nProcessing flight {flight}!')

    # Setup variables
    if path == None:
        path = 'C:\\Documents\\Git 2\\output\\runs\\'

    # Setup variables
    Dir = [i for i in os.listdir(path) if 'meta-analysis' in i][0]
    file_path = os.path.join(path, Dir)
    file = pd.read_excel(file_path)
    shape = np.array(file.shape)
    index = file.index
    columns = file.columns
    palette = ['r', 'y', 'b', 'g']
    timewindows = list([])

    # Find number of sibling aircraft from the file
    for i in range(shape[0]):
        # print(f'iterator is {i} and compared with {file[:].iloc[i][0]}')
        if file[:].iloc[i][0] == i:
            continue
        else:
            N = i
            break

    # Find name and number of schedules
    number_of_schedules = int(float((shape[1] - 1) / 13))
    schedules = list()
    for i in range(number_of_schedules):
        schedules.append(file[:].iloc[N][i * 13 + 2])

    # Get the delays
    delays = list()
    for i in range(N):
        delays.append(file[:].iloc[i][1])

    print(f'The number of sibling aircraft is: {N}')
    print(f'The delays of the aircraft are: \n{delays}')
    # print(f'The names of the schedules are: {schedules}')

    # Create a dataset for every schedule
    selection = list()
    for i in range(int(np.floor(shape[1] / 13))):
        selection.append(columns[i * 13 + 2])
        selection.append(columns[i * 13 + 7])
        selection.append(columns[i * 13 + 10])

    # print(f'Plotting flight {flight}!')
    w, h, l = len(schedules), 2, N
    data = [[[0 for x in range(w)] for y in range(h)] for z in range(l)]
    a = file[columns[2]].str.split('-', n=1, expand=True)
    subfile = file[selection].iloc[:][flight == a[0]]
    subfile2 = file.iloc[:][flight == a[0]]
    subfile.reset_index(drop=True)
    subcolumns = subfile.columns

    # Store the data on the right places
    for i, schedule in enumerate(schedules):
        data[i][0][:] = subfile[subcolumns[i * 3 + 1]].iloc[:]
        data[i][1][:] = subfile[subcolumns[i * 3 + 2]].iloc[:]
        deltatime = datetime.datetime.strptime(subfile2.iloc[2, 5+13*i], '%H:%M:%S') - \
                    datetime.datetime.strptime(subfile2.iloc[2, 4+13*i], '%H:%M:%S')
        timewindows.append(int(deltatime.total_seconds() / 60))
    # print(f'The time-windows are: {timewindows}')

    ind = np.arange(N)  # the x locations for the groups
    width = 0.9 / len(schedules)  # the width of the bars
    legendlist = list()
    d = {}
    fig, ax = plt.subplots(figsize=(12, 3))
    fuel_max = 0
    fuel_min = 10**6
    for i in range(number_of_schedules):
        fuel_max = max(fuel_max, max(data[i][:][0]))
        fuel_min = min(fuel_min, min(data[i][:][0]))
        d[f'rects{i}'] = ax.bar(ind + width * i, data[i][:][0], width, color=palette[i], yerr=data[i][:][1])
        holder = f' - {timewindows[i]}'
        legendlist.append(subfile[subcolumns[3 * i]].str.split('-', n=1, expand=True)[1].iloc[0] + holder)
    print(f'The legend becomes: \n{legendlist}')

    fuel_max = np.ceil(fuel_max/(10**(len(str(int(fuel_min)))-1)))*(10**(len(str(int(fuel_min)))-1))
    fuel_min = np.floor(fuel_min/(10**(len(str(int(fuel_min)))-1)))*(10**(len(str(int(fuel_min)))-1))
    if ylimits:
        fuel_min = ylimits[0]
        fuel_max = ylimits[1]
        if len(ylimits) < 3:
            ylimits.append(1.02)

    # print(f'Fuel max is {fuel_max}')
    # print(f'Fuel min is {fuel_min}')

    # add some text for labels, title and axes ticks
    ax.set_ylabel('Fuel Consumption [kg]', fontsize=12)
    ax.set_xlabel('Delay [seconds]', fontsize=12)
    # ax.set_title(f'Fuel Consumption by Delay and Schedule for Flight {flight}', fontsize=16)
    ax.set_xticks(ind + width * (number_of_schedules - 1) / 2)
    ax.set_xticklabels((delays))

    if len(schedules) == 3:
        ax.legend((d["rects0"][0], d["rects1"][0], d["rects2"][0]), legendlist,
                  loc='upper left', title='Schedule - TW [min]', framealpha=0.1)
    else:
        ax.legend((d["rects0"][0], d["rects1"][0], d["rects2"][0], d["rects3"][0]), legendlist,
                  loc='upper left', title='Schedule - TW [min]', framealpha=0.1)

    if ylimits:
        for i in range(len(schedules)):
            autolabel(d["rects" + str(i)], ax, True, 8, ylimits[2])
    else:
        for i in range(len(schedules)):
            autolabel(d["rects" + str(i)], ax)
    plt.ylim(fuel_min, fuel_max)
    plt.gcf().subplots_adjust(bottom=0.15)
    plt.tight_layout()
    plt.show()
    return legendlist

# Create a plot for the KPI of reaching a TW per flight
# wp_sel: skip these waypoints in the plots
def TWperformance(flights_input, path=None, wp_sel=False):
    # Setup variables
    if path == None:
        path = 'C:\\Documents\\Git 2\\output\\runs\\analysis'
    else:
        path = os.path.join(path, 'analysis')

    # Calculate KPI per flight with matched/violated waypoint per waypoint
    # Firstly see if the waypoint has been matched or violated early/late
    flights = list()
    flights.extend(flights_input)
    kickoff = False
    custom_order = ["min",  "det",   "prob",  "inf"]
    if wp_sel:
        for i, j in enumerate(wp_sel):
            wp_sel[i] = j - 1
        test_x = len(range(min(wp_sel), max(wp_sel)))
    else:
        test_x = 0
    for i_flight, flight in enumerate(flights):
        if i_flight > 0:
            break
        print(f'\nProcessing flight {flight}!')

        # Setup variables
        holder = list()
        schedules = list()
        Dir = [i for i in os.listdir(path) if flight in i]
        holder.extend(set([i.split(' ')[1] for i in Dir]))
        for i in custom_order:
            if i in holder:
                schedules.append(i)

        for i_schedule, schedule in enumerate(schedules):
            print(f'Processing schedule {schedule}')
            dir = [i for i in Dir if schedule in i]

            # Find number of delays
            file_path = os.path.join(path, dir[0])
            file = pd.read_excel(file_path)
            nDelays = 0
            for i in file.iloc[3:, 1]:
                if type(i) == int:
                    nDelays += 1
                else:
                    break
            if kickoff:
                if nDelays != nDelays2:
                    print(f'Warning! The set of delays is not consistent over the schedules.')
                    print(f'The new set of delays consists of {nDelays} delays and the previous set has {nDelays2} delays.')

            # Iterate over the files
            for File in dir:
                if '$' in File:
                    continue
                file_path = os.path.join(path, File)
                file = pd.read_excel(file_path, sheet_name=None)

                for sheetname, sheet in file.items():
                    # Find number of waypoints
                    for i in sheet.iloc[0][7:]:
                        if 'nWP' in str(i):
                            nWP = int(i[6:])-1

                    if not kickoff:
                        w, h, l, m = len(schedules), nWP-test_x, nDelays, 3
                        KPI = np.array([[[[0 for x in range(w)] for y in range(h)] for z in range(l)] for z2 in range(m)])
                        column_names = [f'Waypoint {x+1}' for x in range(nWP)]
                        delays = list(sheet.iloc[3:3+nDelays, 1])
                        print(f'The set of delays are: \n{delays}')
                        print(f'The shape of the KPI-matrix is {KPI.shape}.\n')
                        nDelays2 = nDelays
                        kickoff = True
                    test_i = 0
                    # So the important columns lie in column 8 to 8+nWP
                    for i in range(8, 8+nWP):
                        if wp_sel:
                            if i >= min(wp_sel)+8 and i < max(wp_sel)+8:
                                test_i += 1
                                continue
                        early = sheet.iloc[1, i]
                        late = sheet.iloc[2, i]
                        minTime = datetime.datetime(100, 1, 1, int(early[-8:-6]), int(early[-5:-3]), int(early[-2:]))
                        maxTime = datetime.datetime(100, 1, 1, int(late[-8:-6]), int(late[-5:-3]), int(late[-2:]))
                        # print(f'\nMin time is: {minTime} | Max time is: {maxTime}')

                        for i_delay, j in enumerate(list(sheet.iloc[3:3+nDelays, i])):
                            # print(f'Waypoint {i-7} has arrival time of {j}')
                            curTime = datetime.datetime(100, 1, 1, int(j[-8:-6]), int(j[-5:-3]), int(j[-2:]))
                            # Store the data on the right places
                            if curTime > maxTime:
                                KPI[2, i_delay, i-8-test_i, i_schedule] += 1
                            elif curTime < minTime:
                                KPI[1, i_delay, i-8-test_i, i_schedule] += 1
                            else:
                                KPI[0, i_delay, i-8-test_i, i_schedule] += 1

    def autolabel(rects, text, xpos='center'):
        ha = {'center': 'center', 'right': 'left', 'left': 'right'}
        offset = {'center': 0, 'right': 1, 'left': -1}
        height = -20
        for rect in rects:
            ax.annotate(f'{text}',
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(offset[xpos] * 3, 3),  # use 3 points offset
                        textcoords="offset points",  # in both directions
                        ha=ha[xpos], va='bottom', rotation=90)#, weight='bold') #, fontsize=10)

    # Plot the KPI
    flights = flights[0] # right now, only one flight can be processed at a time
    print(f'\nPlotting {len(schedules)} schedules for Flight {flight}!')
    print(f'These are: {schedules}')

    palette = ['r', 'g', 'y']
    # KPI = np.delete(KPI, 0, [np.arange(2, 7)])
    if wp_sel:
        i = max(wp_sel)-1
        while i >= min(wp_sel):
            column_names.pop(i)
            i += -1
        print(f'Selected columns are: {column_names}')
    for i_schedule, schedule in enumerate(schedules):
        print(f'Plotting schedule [{schedule}]!')
        d = {}
        for i in range(KPI.shape[1]):
            d[f'{alphabet[i*3  ]}']   = KPI[2, i, :, i_schedule] *2
            d[f'{alphabet[i*3+1]}']   = KPI[0, i, :, i_schedule] *2 #+ KPI[2,i,:,tlk]
            d[f'{alphabet[i*3+2]}']   = KPI[1, i, :, i_schedule] *2 #+ KPI[2,i,:,tlk] + KPI[0,i,:,tlk]

        df2 = pd.DataFrame(d)
        width_val = list([0.1 for x in range(int(df2.shape[1] / 3))])
        width_val2 = list([0 for x in range(int(df2.shape[1] / 3))])
        for i in range(int(df2.shape[1] / 6)):
            width_val2[i]    = np.round(-0.4+0.1*i, 1)
            width_val2[-1-i] = np.round(0.3-0.1*i, 1)

        fig, ax = plt.subplots(figsize=(5, 3.5))

        for i in range(int(df2.shape[1]/3)):
            d[f'{alphabet[i*3]}{alphabet[i*3+1]}{alphabet[i*3+2]}_bar_list']   = \
                    [plt.bar(np.arange(KPI.shape[2])+width_val2[i], d[f'{alphabet[i*3+2]}'], #eval(f'df2.{alphabet[i*3+2]}'),
                             align='edge', width=width_val[i], color=palette[2], hatch='//',
                             bottom=d[f'{alphabet[i*3+1]}']+d[f'{alphabet[i*3]}']),
                     plt.bar(np.arange(KPI.shape[2])+width_val2[i], d[f'{alphabet[i*3+1]}'], #eval(f'df2.{alphabet[i*3+1]}'),
                             align='edge', width=width_val[i], color=palette[1], #hatch='\\\\',
                             bottom=d[f'{alphabet[i*3]}']),
                     plt.bar(np.arange(KPI.shape[2])+width_val2[i], d[f'{alphabet[i*3]}'], #eval(f'df2.{alphabet[i*3]}'),
                             align='edge', width=width_val[i], color=palette[0], hatch='..')]
            autolabel(d[f'{alphabet[i*3]}{alphabet[i*3+1]}{alphabet[i*3+2]}_bar_list'][0], delays[i])

        # ax.set_title(f'Punctuality for Flight {flights} - [{schedule.upper()}]', fontsize=16)
        print(f'Punctuality for Flight {flights} - [{schedule.upper()}]')
        ax.set_ylabel('Arrival punctuality [%]', fontsize=14)
        ax.set_xlabel('Delay [seconds] by waypoint [#]', fontsize=14)
        ax.legend(['Too Early', 'On Time', 'Too Late'], loc='upper right')
        plt.xticks(np.arange(KPI.shape[2]), column_names, fontsize=14)
        plt.yticks(np.arange(0, 101, 10))
        plt.ylim(-20, 100)
        plt.tight_layout()
        plt.show()
    return

# Secondly combine the results of the matching and create a KPI
# Score = matched WP * weight factor (distance WP / total distance) * # of matches / (
# Calculate KPI per schedule with overall matched/violated waypoints
# First get percentages to weigh every waypoint
#   ~ use infinite case?
#   ~ should be distance based (?)
#       -> get coordinates
#       -> calculate distances per waypoint and total distance
#       -> divide the distance between waypoints by the total and use it as weightfactors
# Secondly see if the waypoint has been matched or violated early/late
# Thirdly combine the results of the matching and create a KPI
# Create a plot for the KPI of reaching a TW per flight
def TWscore(flights=None, path=None, legend=None, flight_limit=10):
    # Setup variables
    if not path:
        path = 'C:\\Documents\\Git 2\\output\\runs\\analysis'
        path2 = 'C:\\Documents\\Git 2\\output\\runs\\xlogs input\\1 min'
    else:
        path2 = os.path.join(path, 'xlogs input\\1 min')
        path = os.path.join(path, 'analysis')

    path2 = os.path.join(path2, os.listdir(path2)[0])

    # Calculate KPI per flight with matched/violated waypoint per waypoint
    # Firstly see if the waypoint has been matched or violated early/late
    if not flights:
        flights = list(set([x.split(' ')[0] for x in os.listdir(path)]))
        flights.sort()
    # else:

    kickoff = False
    custom_order = ["min",  "det",   "prob",  "inf"]
    palette = ['r', 'y', 'b', 'g']
    print(f'\nNumber of flights is {len(flights)} / {flight_limit}')
    flight_limit -= 1

    for i_flight, flight in enumerate(flights):
        if i_flight > flight_limit:
            break
        print(f'\nProcessing flight {flight}!')
        kickoff_2 = False

        # Setup variables
        Dir = [i for i in os.listdir(path) if flight in i]
        if not kickoff:
            holder = list()
            schedules = list()
            holder.extend(set([i.split(' ')[1] for i in Dir]))
            for i in custom_order:
                if i in holder:
                    schedules.append(i)

        # Calculate weight factors of the waypoints
        # This should be done per flight and only once for all the schedules
        # Find list of waypoints with coordinates
        log = list([line for line in open(path2, 'r') if f'DEFWPT {flight}-min-0-' in line])
        log.append(log.pop(1))
        WeightFactors = list()
        for i in range(len(log)-2):
            j = log[i].split(' ')
            k = log[i+1].split(' ')
            a = float(j[2].strip(","))
            b = float(j[3].strip(","))
            c = float(k[2].strip(","))
            d = float(k[3].strip(","))
            WeightFactors.append(dist(a, b, c, d)[1])
        print(f'Total distance for trajectory {flight} is {round(sum(WeightFactors),2)}')
        WeightFactors = WeightFactors/sum(WeightFactors)

        for i_schedule, schedule in enumerate(schedules):
            # print(f'Processing schedule {schedule}')
            dir = [i for i in Dir if schedule in i]

            # Find number of delays
            file_path = os.path.join(path, dir[0])
            file = pd.read_excel(file_path)
            nDelays = 0
            for i in file.iloc[3:, 1]:
                if type(i) == int:
                    nDelays += 1
                else:
                    break
            if kickoff:
                if nDelays != nDelays2:
                    print(f'Warning! The set of delays is not consistent over the schedules.')
                    print(bcolors.WARNING + f'The new set of delays consists of {nDelays} '
                            f'delays and the previous set has {nDelays2} delays.' + bcolors.ENDC)

            # Iterate over the files
            for File in dir:
                if '$' in File:
                    continue
                file_path = os.path.join(path, File)
                file = pd.read_excel(file_path, sheet_name=None)

                for sheetname, sheet in file.items():
                    # Find number of waypoints
                    for i in sheet.iloc[0][7:]:
                        if 'nWP' in str(i):
                            nWP = int(i[6:])-1

                    if not kickoff:
                        w, h = len(flights), len(schedules)
                        KPI_score = pd.DataFrame(data=flights, columns=['Flights']).set_index('Flights') #[[0 for x in range(w)] for y in range(h)])
                        for i in schedules:
                            KPI_score[i] = np.NAN
                        # print(f'The shape of the KPI_score-matrix is {KPI_score.shape}.\n')
                        kickoff = True

                    if not kickoff_2:
                        w, h, l, m = len(schedules), nWP, nDelays, 3
                        KPI = np.array(
                            [[[[0 for x in range(w)] for y in range(h)] for z in range(l)] for z2 in range(m)])
                        # column_names = [f'Waypoint {x + 1}' for x in range(nWP)]
                        delays = list(sheet.iloc[3:3 + nDelays, 1])
                        print(f'The set of delays are: \n{delays}')
                        # print(f'The shape of the KPI-matrix is {KPI.shape}.')
                        nDelays2 = nDelays
                        kickoff_2 = True

                    # So the important columns lie in column 8 to 8+nWP
                    for i in range(8, 8+nWP):
                        early = sheet.iloc[1, i]
                        late = sheet.iloc[2, i]
                        minTime = datetime.datetime(100, 1, 1, int(early[-8:-6]), int(early[-5:-3]), int(early[-2:]))
                        maxTime = datetime.datetime(100, 1, 1, int(late[-8:-6]), int(late[-5:-3]), int(late[-2:]))
                        # print(f'\nMin time is: {minTime} | Max time is: {maxTime}')

                        for i_delay, j in enumerate(list(sheet.iloc[3:3+nDelays, i])):
                            # print(f'Waypoint {i-7} has arrival time of {j}')
                            curTime = datetime.datetime(100, 1, 1, int(j[-8:-6]), int(j[-5:-3]), int(j[-2:]))
                            # Store the data on the right places
                            if curTime > maxTime:
                                # print('Too late')
                                KPI[2, i_delay, i-8, i_schedule] += 1
                            elif curTime < minTime:
                                KPI[1, i_delay, i-8, i_schedule] += 1
                                # print('Too early')
                            else:
                                KPI[0, i_delay, i-8, i_schedule] += 1
                                # print('On Time !')

            # Multiply the weight factors with the KPI matrix to get a score and store it in KPI_score
            print(f'Saving {flight}-{schedule} in the KPI_score-matrix!')
            for i in range(int(KPI.shape[3])):
                KPI_score.loc[flight, schedules[i]] = np.round(sum(np.sum(KPI[0, :, :, i], axis=0) * 2 / KPI.shape[1] * WeightFactors), 2)

    # Create a table for the KPI of reaching TW per flight
    # one grade/percentage/stat per flight of a schedule
    # take distance into acocunt? If only compared between schedules, not necessary?
    KPI_score.dropna(axis=0, how='any', thresh=2, subset=None, inplace=True)
    print('\n', KPI_score) # Table

    # If only one flight, combine the stats in 1 plot and return
    if i_flight < 1:
        fig, ax = plt.subplots(figsize=(6, 3.5))
        # KPI_score.plot(kind='bar', color='b')
        for i in range(len(KPI_score.columns)):
            # print(KPI_score.iloc[0, i])
            bar = ax.bar(i+1, KPI_score.iloc[0, i])
            autolabel(bar, ax, True, 12, 1.01)
        ax.set_ylabel('Aggregated arrival punctuality [%]', fontsize=12)
        ax.set_xlabel('Schedules - TW [min]', fontsize=12)
        # fig.suptitle(f'Punctuality per Schedule for Flight {flights[0]}',
        #              x=.5, y=.94, fontsize=16)
        # ax.set_suptitle(f'Punctuality per Schedule for Flight {flights[0]}', fontsize=16)
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.xaxis.set_major_locator(MaxNLocator(10))
        ax.set_xticks(range(1, len(KPI_score.columns)+1))
        if legend:
            ax.set_xticklabels(legend, fontsize=10)
        else:
            ax.set_xticklabels(KPI_score.columns.str.upper(), fontsize=10)
        if len(KPI_score.columns) > 3:
            ax.set_xlim(0, 5)
        else:
            ax.set_xlim(0, 4)
        ax.set_ylim(0, 100)
        plt.gcf().subplots_adjust(bottom=0.15)
        plt.show()
        return

    # Create a plot per schedule with those percentages/grades from above for all flights
    # It gives info on #flights that stay within certain scoreranges
    # See the picture. Bars per 10% range for a schedule.
    KPI_bars = np.array([[0 for x in range(10)] for y in range(len(KPI_score.columns))])
    for i2, i in enumerate(KPI_score.columns):
        for j2, j in enumerate(range(0, 100, 10)):
            holder = np.where((KPI_score[i].values >= j) & (KPI_score[i].values < j+10), 1, 0)
            # print(f'j = {j} and j+10 = {j+10} => holder = {holder} => sum = {sum(holder)}')
            KPI_bars[i2, j2] = sum(holder)
        # print(f'schedule = {i} and KPI_bars = {KPI_bars}')

    # Setup for the plot
    if KPI_bars.shape[0] % 2 == 0:
        i_1 = 2
        i_2 = 2
    else:
        i_1 = 2
        i_2 = 1
    x = range(5, 105, 10)
    KPI_max = math.ceil(int(np.max(KPI_bars))/5) * 5

    # Create the plot
    fig = plt.figure()
    # fig.suptitle('Flights vs Punctuality per schedule', x=.5, y=.95, fontsize=16)
    fig.set_figheight(6)
    fig.set_figwidth(9)
    for i2, i in enumerate(range(KPI_bars.shape[0])):
        ax = fig.add_subplot(i_1, i_2, i2+1)
        ax.set_title(f'Distribution of {(KPI_score.columns[i2]).upper()}-schedule', fontsize=14)
        bar = ax.bar(x, KPI_bars[i2, :], width=8)
        autolabel(bar, ax, True, 10)
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.tick_params(axis='y', pad=-3)
        # ax.tick_params(axis='x')
        ax.xaxis.set_major_locator(MaxNLocator(10))
        xlim = range(0, 101, 10)
        ax.set_xticklabels(xlim, fontsize=10, verticalalignment='bottom')

        ax.set_xlim(0, 100)
        ax.set_ylim(0, KPI_max)

        if i2+1 > (KPI_bars.shape[0] / 2):
            ax.set_xlabel('Aggregated arrival punctuality [%]', fontsize=14)
        if i2 == (KPI_bars.shape[0] / 2) or i2 < 1:
            ax.set_ylabel('Flights [#]', fontsize=14)
    plt.gcf().subplots_adjust(bottom=0.15)
    plt.subplots_adjust(wspace=0.10, hspace=0.20)

    plt.show()

    return

# Create a plot for number of speed changes per flight --> present these in a plot flights vs changes ?
# Do this for every schedule or smthing? what does this tell the user?
def speedchanges(flights=None, path=None, font=8):
    # Setup variables
    if path == None:
        path = 'C:\\Documents\\Git 2\\output\\runs\\analysis'
    else:
        path = os.path.join(path, 'analysis')

    if not flights:
        flights = list(set([x.split(' ')[0] for x in os.listdir(path)]))
        flights = [x for x in flights if "$" not in x]
        flights.sort()
        nFlights = len(flights)
        print(f'Number of flights is {nFlights}')
    else:
        flights.append('stop')
        nFlights = 1
    kickoff = False
    custom_order = ["min", "det", "prob", "inf"]

    for i_flight, flight in enumerate(flights):
        # Stop after # of flights or at second flight if only one has been called
        if i_flight > 10 or 'stop' in flight:
            break
        print(f'\nProcessing flight {flight}!')
        kickoff_2 = False
        timewindows = list([])

        # Setup variables
        Dir = [i for i in os.listdir(path) if flight in i]
        if not kickoff:
            # Find order of schedules (either 3 or 4 schedules)
            holder = list()
            schedules = list()
            holder.extend(set([i.split(' ')[1] for i in Dir]))
            for i in custom_order:
                if i in holder:
                    schedules.append(i)

        for i_schedule, schedule in enumerate(schedules):
            # print(f'Processing schedule {schedule}')
            kickoff_3 = True
            dir = [i for i in Dir if schedule in i]

            # Find number of delays
            file_path = os.path.join(path, dir[0])
            file = pd.read_excel(file_path)
            nDelays = 0
            for i in file.iloc[3:, 1]:
                if type(i) == int:
                    nDelays += 1
                else:
                    break

            # Keep checking the number of delays as a consistency check
            if kickoff:
                if nDelays != nDelays2:
                    print(bcolors.WARNING + f'Warning! The set of delays is not consistent over the schedules.')
                    print(f'The new set of delays consists of {nDelays} '
                            f'delays and the previous set has {nDelays2} delays.' + bcolors.ENDC)
                    input('Press a button to continue')

            # Iterate over the files
            for File in dir:
                if '$' in File:
                    continue
                file_path = os.path.join(path, File)
                file = pd.read_excel(file_path, sheet_name=None)

                for sheetname, sheet in file.items():
                    if not kickoff:
                        nFlights = nFlights * nDelays
                        delays = list(sheet.iloc[3:3 + nDelays, 1])
                        flights_names = [None] * nFlights
                        if 'stop' in flights:
                            flights_names[0:nDelays] = [f'{flights[0]}-{x}' for x in delays]
                        else:
                            for j, i in enumerate(delays):
                                flights_names[j::nDelays] = [x + f'-{i}' for x in flights]
                        KPI_score = pd.DataFrame(data=flights_names, columns=['Flights']).set_index('Flights')
                        for i in schedules:
                            KPI_score[f'{i}-pos'] = 0
                            KPI_score[f'{i}-neg'] = 0
                        print(f'The shape of the KPI_score-matrix is {KPI_score.shape}.')
                        start_cell = (nDelays+4) * 2
                        kickoff = True

                    if not kickoff_2:
                        delays = list(sheet.iloc[3:3 + nDelays, 1])
                        print(f'The set of delays are: \n{delays}')
                        nDelays2 = nDelays
                        kickoff_2 = True

                    if kickoff_3:
                        deltatime = datetime.datetime.strptime(sheet.iloc[2, 4], '%H:%M:%S') - \
                                    datetime.datetime.strptime(sheet.iloc[1, 4], '%H:%M:%S')
                        timewindows.append(int(deltatime.total_seconds() / 60))
                        kickoff_3 = False

                    sheet = sheet.iloc[start_cell:]
                    sheet.reset_index(inplace=True)
                    sheet.drop(['index', 'Waypoint Analysis'], axis=1, inplace=True)
                    sheet.columns = sheet.iloc[0]
                    sheet.drop([0], inplace=True)
                    # print('__________________________')
                    # print(f'Sheet : {sheetname}')

                    # So the important columns lie in column (0, 1, 2)*n
                    for i, delay in enumerate(delays):

                        nInputs = int(sheet.columns[i*3+1].split(' ')[1])
                        # print(f'\nDelay : {delay}')
                        # print(f'Inputs : {nInputs}')
                        if nInputs == 0:
                            continue
                        initial_speed = float(sheet.columns[i*3+2].split(' ')[3])
                        inputs = np.array(sheet.iloc[1:nInputs+1, i*3+2].values.astype(float))
                        inputs2 = np.copy(inputs)
                        inputs2[0] = float(sheet.columns[i*3+2].split(' ')[3])
                        inputs2[1:] = inputs[:-1]
                        # print(f'Decelerations : {np.sum([inputs2 < inputs]*1)}')
                        # print(f'Accelerations : {np.sum([inputs2 > inputs]*1)}')
                        KPI_score.loc[f'{flight}-{delay}', f'{schedule}-pos'] += np.sum([inputs2 < inputs]*1)
                        KPI_score.loc[f'{flight}-{delay}', f'{schedule}-neg'] += np.sum([inputs2 > inputs]*1)
        # print(f'\nStarting speed : V = {initial_speed} Mach')
        # print(KPI_score.iloc[i_flight*nDelays:(i_flight+1)*nDelays, :])
    KPI_score = KPI_score/50

    # Setup for the plot
    if KPI_score.shape[1] % 8 == 0:
        i_1 = 2
        i_2 = 2
    else:
        i_1 = 2
        i_2 = 2

    if max(KPI_score.max()) >= 5:
        KPI_max = math.ceil(max(KPI_score.max())/5) * 5
    else:
        KPI_max = math.ceil(max(KPI_score.max()))+1
    ind = np.arange(nDelays)  # the x locations for the groups
    width = 0.8 / 2  # the width of the bars
    legendlist = list(['acceleration', 'deceleration'])
    palette = ['r', 'y', 'b', 'g']
    # print(f'KPI_score shape is {KPI_score.shape}')
    # print(f'Delays is {delays}')

    # Create the plot
    fig = plt.figure()
    # fig.suptitle(f'Speed input vs Delay per schedule for Flight {flights[0]}',
    #              x=.5, y=.94, fontsize=16)
    rcParams['axes.titlepad'] = 5
    fig.set_figheight(6)
    fig.set_figwidth(8)
    # gs1 = gridspec.GridSpec(4, 4)
    # gs1.update(wspace=-20, hspace=-1)
    print(int(KPI_score.shape[1]/2))
    for i2, i in enumerate(range(int(KPI_score.shape[1]/2))):
        d = {}
        ax = fig.add_subplot(i_1, i_2, i2 + 1)
        d[f'rects1'] = ax.bar(ind - width,  KPI_score.loc[:, f'{schedules[i]}-pos'], width, color=palette[2])
        d[f'rects2'] = ax.bar(ind,          KPI_score.loc[:, f'{schedules[i]}-neg'], width, color=palette[3])

        ax.set_title(f'{schedules[i2].upper()} - TW: {timewindows[i2]} [min]', fontsize=12)
        ax.set_xticks(ind-width/2)
        ax.set_xticklabels(delays, fontsize=10, verticalalignment='bottom')
        # plt.labelpad = 10
        if i2 < 5:
            ax.legend((d["rects1"][0], d["rects2"][0]), legendlist, loc='upper right')
        else:
            ax.legend((d["rects1"][0], d["rects2"][0]), legendlist, loc='upper left')

        autolabel(d["rects1"], ax, False, font)
        autolabel(d["rects2"], ax, False, font)

        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        ax.tick_params(axis='y', pad=-3)
        ax.set_ylim(0, KPI_max)
        if i2+1 > (KPI_score.shape[0] / 4):
            ax.set_xlabel('Delay [seconds]')
        if i2 % 2 == 0:
            ax.set_ylabel('Average number of speed changes [#]')
        # ax.xaxis.labelpad = -5
        # print(ax.get_xticks())
    plt.subplots_adjust(wspace=0.10, hspace=0.17)
    plt.show()
    return