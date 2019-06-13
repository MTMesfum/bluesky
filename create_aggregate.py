import os
import pandas as pd
import numpy as np

path = "C:\Documents\Git 2\output\\runs ----"
Dir = os.listdir(path)
to_save = pd.DataFrame()

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

for i, dir in enumerate(Dir):
    # if 'min' in dir:
    #     print('skipped ', dir)
    #     continue
    if 'skip' in dir:
        print('skipped ', dir)
        continue
    if 'put' in dir:
        print('skipped ', dir)
        continue
    # if 'det' in dir:
    #     print('skipped ', dir)
    #     continue

    Files = os.listdir(path + '\\' + dir)
    l = 1

    for j, file in enumerate(Files):
        if '~$' in file:
            continue

        if j == 0:
            apple = pd.read_excel(path + '\\' + dir + '\\' + file, index_col=None, header=None)
            banana = list(range(0, len(apple.columns)))
            to_pop = list([0, 1, 2, 3, 5, 6, 7])
            to_pop.reverse()
            for k in to_pop:
                banana.pop(k)
            apple = apple.drop(columns=banana, axis=1)
            # print(apple)
            if i == 0:
                to_save = apple.drop(columns=list([0, 1, 5, 6, 7]))
            else:
                to_save = pd.concat([to_save, apple[3]], axis=1)
            apple = apple.drop(columns=list([0, 1, 2, 3, 5]))
            # apple[6] = pd.to_timedelta(apple[6], unit='s')
            # else:
            #     apple = apple.drop(columns=banana, axis=1)
                    # pd.concat([to_save, apple], axis=1)
                # to_save = to_save.append(apple(columns=[2]))
                # print(to_save)
                # to_save = to_save.append(apple[3])
            continue

        apple2 = pd.read_excel(path + '\\' + dir + '\\' + file, index_col=None, header=None)
        banana = list(range(8, len(apple2.columns)))
        # apple2 = apple2.drop(columns=banana, axis=1)
        # print(dir, file)
        # print(pd.to_timedelta(apple[6]))
        # print(apple2)
        # pd.to_timedelta(df.hour_in + ':00', errors='coerce')
        # print(apple[6])
        # print(apple2[6])
        apple[6] = pd.to_timedelta(apple[6], unit='s') + pd.to_timedelta(apple2[6], unit='s')
        # print(file, apple[6])
        apple[7] = apple[7] + apple2[7]
        l += 1

    # print(apple)
    apple[6] = (apple[6] / l).dt.round('1s')
    apple[7] = round(apple[7] / l, 2)
    # print(apple[6])
    # to_save = to_save.append(apple(columns=[6,7]))
    # citrus = list(range(0, 6))
    # citrus.pop(4)
    # apple =
    # print(apple)
    # print()
    # to_save = to_save.append(apple.drop(columns=citrus, axis=1))
    to_save = pd.concat([to_save, apple], axis=1)
    del apple, apple2
# print(to_save)
filename = path + '\\' + 'analysis{}.xlsx'.format(19)
with open(filename, 'wb') as f:
    # df.set_index(['Name', 'Std'])
    to_save.columns = ([    'Delay', 'Min', 'Arrival 1', 'Fuel Con 1',
                            'Det', 'Arrival 2', 'Fuel Con 2',
                            'Prob', 'Arrival 3', 'Fuel Con 3',
                            'Inf', 'Arrival Time', 'Fuel Consumed'])
    to_save[0:8].to_excel(f)
    # print(to_save)
                                        #
# with open(path + '\\' + 'analysis{}.xlsx'.format(18), 'wb') as f:

for k, i in enumerate(range(1, 6)):
    j = i*8
    l = k + j + 1
    append_df_to_excel(filename, to_save[j:j+8], 'Sheet1', l)
        # to_save[j:j+8].to_excel(f, startrow=l) #datetime_format='hh:mm:ss.000')
        # df.to_excel(writer, startrow=len(df) + 2, index=False)
    # exit()
# print(to_save)
exit()
print(files)


